"""The write-tool half of the in-house repo Developer, split out of `adapters/repo_developer.py`
along the tool-vs-persona boundary (docs/15 mega-refactor): `RepoWriteTools` (the surface-gated
write/edit/delete/declare_stages tool provider whose writes are COLLECTED into `self.files`, not
applied — the orchestrator materializes them into the node workdir), the stage-input validators
it shares with the persona's STAGES phase (`_missing_stage_input_paths` /
`_missing_paths_feedback` and their helpers), and the `_xlsx_to_markdown` results renderer the
persona's results context uses.

The persona half (`LLMRepoDeveloper`, `LLMOnboarder` and the prompt constants) stays in
`repo_developer.py`, which re-imports these names — so `looplab.adapters.repo_developer` (and
`repo_task`'s own re-export chain on top of it) keeps exporting them, and this module needs
nothing from `repo_developer` at import time (no cycle)."""
from __future__ import annotations

import json
import os
import posixpath
import re
from typing import Optional

from looplab.tools.edit_match import apply_search_replace
from looplab.tools.patch import SurfacePolicy

# Absolute paths to INPUT data files referenced in a stage command. Only clear INPUT-data extensions
# (a checkpoint .ckpt/.pt an earlier stage WRITES is deliberately excluded, and relative paths resolve
# to mounts at eval time so are left to the eval). Used to catch the #1 real failure: a train stage
# pointing at a hallucinated argparse-default `.pck` that isn't on this machine.
# The leading `/` must be a TRUE absolute-path boundary — the negative lookbehind rejects a `/` that is
# part of a relative `./dir/...` or `a/b/...` (those resolve to mounts/workdir at eval time, not here).
_INPUT_DATA_RE = re.compile(
    r"(?<![\w./~])/[^\s\"',:]+\.(?:pck|parquet|csv|tsv|npy|npz|pkl|arrow|jsonl|feather|h5|hdf5)")

# A flag names an OUTPUT (a path the stage WRITES) when it contains one of these hints — matched on the
# de-dashed flag so ANY spelling works (`--outdir`, `--export-dir`, `--save_to`, `--dest`, `--dump`,
# `--write-to`), replacing the old hardcoded list that missed `--outdir`/`--export-dir`/`--dump`.
_OUTPUT_HINT_RE = re.compile(r"(out|save|dest|export|dump|writ)", re.I)


def _stage_output_values(cmd) -> list[str]:
    """Path values a single stage WRITES: the token after an output-ish flag, or the RHS of
    `--outflag=VAL`. These are pipeline intermediates a LATER stage reads (and are this stage's OWN
    outputs, not its inputs), so they must not be flagged as "missing input" at declare time."""
    out: list[str] = []
    for i, tok in enumerate(cmd):
        if not isinstance(tok, str):
            continue
        if tok.startswith("-") and "=" in tok:
            flag, val = tok.split("=", 1)
            if _OUTPUT_HINT_RE.search(flag.lstrip("-")):
                out.append(val)
        elif tok.startswith("-") and _OUTPUT_HINT_RE.search(tok.lstrip("-")) \
                and i + 1 < len(cmd) and isinstance(cmd[i + 1], str):
            out.append(cmd[i + 1])
    return out


def _covered_by(m: str, produced: list) -> bool:
    """True when absolute path `m` equals, or lives under, a path some stage PRODUCES (exact match, or
    `m` under a produced DIRECTORY like `--outdir /x/prep` covering `/x/prep/train.npy`)."""
    # A falsy produced value is SKIPPED, not treated as a covering directory. A stage command like
    # `--output=` (or the `--save ''` space form) makes `_stage_output_values` yield "", which lands in
    # `produced`; then `m.startswith("".rstrip("/") + "/")` is `m.startswith("/")` — true for EVERY
    # absolute path. One empty output value therefore reported every hallucinated `/…/train.pck` input
    # as covered, in that stage and in all following ones (`produced.extend` keeps the ""), and
    # `_missing_stage_input_paths` returned [] — silently disabling the whole guard.
    for v in produced:
        if v and (v == m or m.startswith(v.rstrip("/") + "/")):
            return True
    return False


def _missing_stage_input_paths(stages) -> list[str]:
    """ABSOLUTE input-data paths referenced in stage commands that DON'T exist on disk — almost always
    a hallucinated default (the recurring failure: a train stage's `--train_dataset /…/train.pck` that
    was copied from the repo's argparse default and isn't here). Absolute paths are location-invariant,
    so a declare-time existence check is sound; relative paths (mounts) and `%params%` are skipped.
    A path an EARLIER stage PRODUCES (or a parent dir of one), or this stage's OWN output, is excluded —
    a valid data_prep→train pipeline's intermediate legitimately doesn't exist yet at declare time. The
    check is stage-ORDER-aware: only outputs of stages at-or-before the reader count, so a read-before-
    write ordering (train reads what a LATER export writes) is still flagged as the FileNotFoundError
    it is."""
    missing: list[str] = []
    produced: list = []                     # output paths of stages processed so far (order-aware)
    for s in (stages or []):
        if not isinstance(s, dict):
            continue
        cmd = [t for t in (s.get("command") or []) if isinstance(t, str)]
        own_outputs = _stage_output_values(cmd)
        known = produced + own_outputs      # this stage's own outputs are not its inputs
        for tok in cmd:
            if "%params%" in tok:
                continue
            for m in _INPUT_DATA_RE.findall(tok):
                if m in missing or _covered_by(m, known):
                    continue
                if not os.path.exists(m):
                    missing.append(m)
        produced.extend(own_outputs)        # available to every stage that FOLLOWS
    return missing


# --- The source-path / manifest COLLISION rule (docs/29 F1c) --------------------------------------
# What may CONTINUE a path token once `<source root>/` has been matched. Deliberately a stop-set
# rather than an allow-set of "path characters": generated content spells these paths inside YAML
# scalars, JSON strings, f-strings and argv lists, and an experiment directory legitimately contains
# `-`, `_`, `.`, `=` and digits. The stop-set is the punctuation that ENDS a path in each of those
# hosts; `*`/`?` stop it too, because a glob is not a path this rule can reason about.
_SRC_PATH_TAIL = re.compile(r"[^\s\"'`,;:()\[\]{}<>|*?\\]*")


def _path_components(p) -> tuple:
    """A path as its ordered, normalized components — the granularity the collision test compares at.

    Component granularity is the whole reason this rule has no false positives. A STRING prefix test
    would fire on `models/rubertlite-20e-v7/last.ckpt` against a declared `models/rubertlite_run/…`
    the moment the two shared the characters `models/rubertlite`, and `runs/rubertlite-dense-retrieval`
    node 36 — the measured 1-in-116 legitimate teacher checkpoint that lives INSIDE the editable root
    — is exactly that shape. Compared as components, `rubertlite-20e-v7` != `rubertlite_run` and the
    two paths merely share a top directory, which is not a collision and never fires.
    """
    p = posixpath.normpath(str(p or "").replace("\\", "/").strip())
    if p in ("", ".", "/"):
        return ()
    return tuple(c for c in p.strip("/").split("/") if c not in ("", "."))


def declared_output_paths(manifest_text: str) -> list[tuple]:
    """`[(stage name, expect.files entry)]` for a staged `looplab_stages.json`, or [].

    Reads the manifest the way `_materialized_stage_list` does — both the wrapped `{"stages": […]}`
    shape `declare_stages` authors and a bare top-level list — but WITHOUT `validate_stages`, on
    purpose: this rule runs while the model is still authoring, and a manifest that would be dropped
    at consume time still states what the agent BELIEVES its pipeline writes, which is the half this
    comparison needs. An unparseable manifest yields [] and the check degrades to silence.

    TOTAL over any JSON document, and that is the contract rather than an implementation detail: the
    manifest is MODEL-AUTHORED (`declare_stages` validates, but a hand-written `looplab_stages.json`
    is an accepted Developer surface — see `repo_task.py`'s stage intake), this runs from `_write`
    on EVERY subsequent write in the session, and `agents/tool_loop.py::_run_tool_call` does not
    contain a tool exception — so a raise here leaves the tool loop as a Developer crash, which
    `orchestrator.py` turns into a node terminal AND a run-level auto-pause. Guarding only
    `json.loads` was not enough: `{"stages": 7}` and a bare `5` are PARSEABLE and not iterable, and
    a non-dict `expect` / a scalar `expect.files` are the same shape one level down. Each guard is a
    type test rather than a wider `except`, so a real bug here still raises.
    """
    try:
        obj = json.loads(manifest_text or "")
    except (ValueError, TypeError):
        return []
    stages = obj.get("stages") if isinstance(obj, dict) else obj
    out: list[tuple] = []
    if not isinstance(stages, (list, tuple)):
        return []                                 # `{"stages": 7}` / a bare scalar document
    for s in stages:
        if not isinstance(s, dict):
            continue
        expect = s.get("expect")
        files = expect.get("files") if isinstance(expect, dict) else None
        if not isinstance(files, (list, tuple)):
            continue                              # `"expect": 7` / `"files": 3`
        for f in files:
            if isinstance(f, str) and f.strip():
                out.append((str(s.get("name") or "?"), f))
    return out


def source_root_targets(content: str, name, root) -> list[tuple]:
    """Every `<root>/<rel>` spelled in `content`, as `(the path as spelled, workdir-relative components)`.

    The translation into the WORKDIR frame is what makes the comparison against `expect.files`
    meaningful at all: a node evaluates in its own materialized copy, where the root editable's tree
    lands at the workspace root and a NAMED editable lands at `wd/<name>` (`engine/workspace.py`).
    So `<root>/vectorsearch/x` is the same place as the manifest's `vectorsearch/x` for the root
    editable, and as `<name>/vectorsearch/x` for a named one — and getting that wrong in a
    multi-editable setup would silently never fire.
    """
    r = str(root or "").replace("\\", "/").rstrip("/")
    # A root of "" or "/" would match every absolute path; a relative root cannot appear as an
    # absolute path in generated code at all. Same guard as `_source_root_paths`.
    if len(r) < 2 or not r.startswith("/"):
        return []
    pre = () if name in ("", ".", None) else _path_components(name)
    out: list[tuple] = []
    for m in re.finditer(re.escape(r + "/"), content):
        rel = _SRC_PATH_TAIL.match(content, m.end()).group(0).rstrip("/.,;:")
        if not rel:
            continue
        out.append((r + "/" + rel, pre + _path_components(rel)))
    return out


def manifest_path_collisions(content: str, manifest_text: str, roots) -> list[tuple]:
    """`[(path as spelled, stage, declared expect.files entry)]` — the F1c collision, decided from
    two artifacts LoopLab already owns and the agent itself authored.

    A collision is: an ABSOLUTE path into an editable repo's SOURCE tree whose workdir-relative
    equivalent is on the same directory chain as a path THIS NODE's manifest declares its own
    pipeline WRITES. That is an unambiguous "this node writes it here and reads it there"
    contradiction — a node runs in its own materialized copy, so a path rooted at the source can
    never name anything this node produced.

    WHY THIS SHAPE AND NOT A BAN ON ABSOLUTE SOURCE PATHS. Measured over every authored working set
    in `runs/` (2,577 `node_created`/`node_repaired` file maps across the corpus, i.e. exactly what
    this function would have been handed): a blanket ban refuses **8 distinct nodes**, of which
    **5 are legitimate** — three `rubert-dr-0807` nodes reading a committed base model at
    `models/converted/e5-small-v1.1.1`, `rubertlite-dense-retrieval` node 36's teacher checkpoint,
    and `rubertlite-dr-unified-v2` node 0, whose config names the repo's own committed default for a
    DIFFERENT experiment than the one it declares. The collision rule refuses **3**, and they are
    exactly the three defective nodes: v2 node 4 and v6 node 4 (both recorded 0.224975 — a human's
    July checkpoint) and v6 node 0 (which trained twice and threw away 2.5 GPU-hours). **8 → 3, and
    0 false positives**, because a legitimate INPUT is never a path the node's own manifest declares
    it WRITES. That property, not a tuned threshold, is what makes the rule safe.

    Both directions of the prefix count, and both occur: v6 node 0's config names the `final/`
    DIRECTORY while its manifest declares `final/model.safetensors` inside it, and v2 node 4's
    config and manifest name the same directory exactly.
    """
    outputs = [(stage, decl, _path_components(decl))
               for stage, decl in declared_output_paths(manifest_text)]
    if not outputs:
        return []                       # no declaration to collide with — degrade to silence
    hits: list[tuple] = []
    for name, root in (roots or []):
        for spelled, cand in source_root_targets(content, name, root):
            for stage, decl, dc in outputs:
                n = min(len(cand), len(dc))
                if n and cand[:n] == dc[:n] and (spelled, stage, decl) not in hits:
                    hits.append((spelled, stage, decl))
    return hits


def collision_feedback(hits: list[tuple], *, where: str = "") -> str:
    """The refusal text. It names BOTH declarations and BOTH ways out, because the model authored
    both and either one may be the one that is wrong — a refusal that names only the path reads as
    "try a different string" and gets answered with a different absolute path.

    Cost of getting this wrong in the other direction is why it is a refusal at the TOOL call rather
    than a note: the same channel already refuses a write that would not compile ("nothing was
    staged"), and the model simply writes again. A note is what shipped for the general case and it
    was measured to be spent — `_source_root_note` fired VERBATIM on `runs/rubertlite-dr-unified-v6`
    node 4's own edit and the node still scored somebody else's model.
    """
    spelled, stage, decl = hits[0]
    return (f"(refused: {('the content you wrote for ' + where) if where else 'this content'} "
            f"hard-codes {spelled} — an ABSOLUTE path into the editable repo's SOURCE tree that "
            f"names the SAME artifact your own stage {stage!r} declares it WRITES to "
            f"{decl!r} (workdir-relative). Those two cannot both be true: this node evaluates in its "
            "OWN materialized copy of the repo, so the source tree can never hold anything this "
            "node's pipeline produced — a `.exists()` on it fails on every node, on every attempt, "
            "forever, and if a file HAPPENS to be there it is somebody else's. Fix ONE of them: "
            "either spell this path RELATIVE to the eval workdir (the same string your `expect` "
            "declares), or, if you really meant a pre-existing input that is not this node's output, "
            "declare it as a `data:`/`references:` mount and stop naming it as your artifact. "
            "Nothing was staged.)")


def _missing_paths_feedback(missing: list[str]) -> str:
    """The actionable bounce message shown to the Developer so it re-declares with a real path.
    Deliberately does NOT tell it to list/inspect the data itself: its scout tools reach ONLY the
    editable repo (mounted inputs materialize in per-node EVAL workdirs it cannot see from here), so
    the old "`list_dir` the actual data" advice just burned the phase's retries on "(path not
    allowed…)" refusals (P13). The authoritative source for a data path is the task/goal/data brief."""
    return ("these data paths in your stage command(s) DO NOT EXIST on this machine: "
            + ", ".join(missing[:5]) + ". Do NOT use the repo's DEFAULT argparse dataset paths — they "
            "are the original author's and aren't here. Take the dataset path from the task/goal/data "
            "brief VERBATIM (mounted inputs appear at ./<name> in the EVAL workdir at run time — your "
            "scout tools cannot list them here) and use it in the stage command, spelled exactly as "
            "given. (If a path is produced by an EARLIER stage, reference it relatively.)")


class RepoWriteTools:
    """Write side of the in-house repo developer (the LLM authors/edits files via tools). Writes are
    COLLECTED into `self.files` (path -> content) rather than applied to disk — the orchestrator
    materializes them into the node workdir as the node's files, surface-gated + protected-skipped
    just like an external coding agent's diff. The SAME gates are enforced here so the model gets
    immediate feedback (a refused write) instead of having the edit silently dropped downstream."""

    def __init__(self, surface, protected, prefixes=None, editables=None,
                 operator_stages: bool = False, data_mounts=None, time_budget=None):
        self.files: dict[str, str] = {}
        self.deleted: list[str] = []
        self._surface = list(surface or [])
        self._protected = set(protected or [])
        self._prefixes = list(prefixes or [])
        # The OPERATOR declared this task's pipeline via `cmd.stages`: the engine runs it verbatim
        # and IGNORES any Developer manifest (_resolve_stages prefers a valid operator list), so
        # declare_stages must REFUSE instead of "succeeding" into a file nobody reads — a repair that
        # "fixed" a stage timeout via the manifest otherwise loops the identical failure to abandon
        # (mega-review P12).
        self._operator_stages = bool(operator_stages)
        # The OPERATOR's per-eval wall-clock ceiling (`command_eval.eval_spec_time_budget`), or None
        # when the task declares no eval spec. A stage `timeout` above it is refused at DECLARATION
        # time — see `stage_budget_refusal`. Passed in rather than derived here because this class
        # holds no task: the repo Developer resolves it ONCE (`_eval_time_budget`) for both the note
        # it states in the prompt and the bound it enforces, so the two cannot name different numbers.
        self._time_budget = time_budget
        # Names of read-only DATA MOUNTS (they sit in the protect list defensively — see
        # RepoTask._protected_names) so a refused write can name the REAL reason: "read-only data
        # mount, write derived data elsewhere" rather than the misleading "the operator owns the eval".
        self._data_mounts = [str(n).rstrip("/") for n in (data_mounts or []) if n]
        # Editable repo roots ({name,path}...) so edit_file can patch a file the node hasn't staged
        # yet: current content = staged overlay first, else the original file on disk.
        self._roots = [(e.get("name") or "", e.get("path")) for e in (editables or []) if e.get("path")]

    def _current(self, p: str):
        """The file's CURRENT content for patching: the staged overlay wins (parent files pre-seeded
        by implement_from, or an earlier write this turn), else the original from an editable root.
        Staged paths are workdir-relative and PREFIXED with the editable's name in multi-editable
        setups (the repo mounts at wd/<name>), so strip the owning prefix before joining its root —
        a bare join would probe <root>/<name>/<file> and read a missing (or wrong) original."""
        if p in self.files:
            return self.files[p]
        from pathlib import Path as _P
        for name, r in self._roots:
            rel = p[len(name) + 1:] if name and name != "." and p.startswith(name + "/") else p
            f = _P(r) / rel
            try:
                if f.is_file():
                    return f.read_text(encoding="utf-8", errors="replace")
            except OSError:
                continue
        return None

    @staticmethod
    def _safe_rel(p: str):
        """Canonicalize to a REPO-RELATIVE path or None. Rejects absolute paths and `..` escapes so
        the agent can only stage files inside the repo it edits — without this, an absolute path like
        `/tmp/x.py` slips past a `**/*.py` surface glob (fnmatch's `*` crosses `/`) and the write is
        silently dropped downstream (it's outside the node workdir)."""
        p = str(p or "").replace("\\", "/").strip()
        # A DRIVE-qualified path (`C:/x.py`, or `C:\\Windows\\x.py` after the separator swap above)
        # passes every POSIX-shaped check below — it neither starts with `/` nor contains `..` — so the
        # write was staged, reported OK to the agent, and then silently discarded by
        # engine/workspace.py's `if wd not in target.parents` fence. Reject it here so the agent gets
        # a real refusal instead of a phantom success. (Checked with PureWindowsPath on every
        # platform: the string came from the model, not from this host's filesystem.)
        from pathlib import PureWindowsPath
        if PureWindowsPath(p).drive:
            return None
        while p.startswith("./"):
            p = p[2:]
        if not p or p.startswith("/") or p.startswith("~") or p == ".." \
                or p.startswith("../") or "/../" in p:
            return None
        return p

    def specs(self) -> list[dict]:
        from looplab.tools._base import fn_spec
        return [
            fn_spec("edit_file",
                     "Edit an EXISTING file with a minimal SEARCH/REPLACE patch — STRONGLY PREFERRED "
                     "over write_file for changing existing code: it is far faster and safer than "
                     "re-generating a whole file. `search` must be copied EXACTLY (including "
                     "whitespace/indentation) from the file's current content and must occur exactly "
                     "once; `replace` is its replacement. Make several small edit_file calls for "
                     "several changes. Use write_file only for NEW files.",
                     {"path": {"type": "string", "description": "repo-relative path"},
                      "search": {"type": "string", "description": "exact existing snippet (unique in the file)"},
                      "replace": {"type": "string", "description": "the replacement snippet"}},
                     ["path", "search", "replace"]),
            fn_spec("write_file",
                     "Create or OVERWRITE a file in the experiment repo you are editing. Provide the "
                     "FULL file content (not a diff, not a shell command). Use this ONLY to author the "
                     "eval entrypoint and code edits — NOT to inspect files. Path is REPO-RELATIVE "
                     "(e.g. test_looplab.py); absolute paths and paths outside the repo are rejected.",
                     {"path": {"type": "string", "description": "repo-relative path, e.g. test_looplab.py"},
                      "content": {"type": "string", "description": "the complete file content"}},
                     ["path", "content"]),
            fn_spec("delete_file",
                     "Delete a file you previously wrote in this experiment (within your surface).",
                     {"path": {"type": "string"}}, ["path"]),
            # The pipeline is AUTHORED in the Developer's dedicated STAGES phase (its `declare_stages`
            # emit) BEFORE implement — but a write session still needs a validated route to FIX the
            # manifest: a repair whose root cause is a bad stage (wrong argv / too-low timeout) has no
            # other way to change it (write_file refuses under the default *.py surface; without this
            # spec every repair repeats the identical stage failure until abandon — mega-review D1).
            fn_spec("declare_stages",
                     "FIX the eval pipeline manifest (looplab_stages.json). The stages were already "
                     "declared in the STAGES phase — call this ONLY when the failure you are fixing is "
                     "IN the pipeline itself (a stage's command/timeout/name is wrong), passing the "
                     "FULL corrected ordered list of preceding stages (the operator's protected `score` "
                     "step stays appended after them). `%params%` in a command injects this node's "
                     "hyperparameters; give a long `train` a generous `timeout` (seconds). It VALIDATES "
                     "the manifest and reports errors back. Do not use it to re-plan working stages."
                     + (" NOTE: THIS task's pipeline is OPERATOR-declared (`cmd.stages`) and runs "
                        "verbatim — this tool will refuse; fix the failing stage's script instead."
                        if self._operator_stages else ""),
                     {"stages": {"type": "array", "description":
                                 "ordered preceding stages, each {name, command:[argv...], timeout?, "
                                 "check?, needs?:[input paths this stage READS], "
                                 "expect?:{files:[output paths this stage WRITES], assert?}}"}},
                     ["stages"]),
        ]

    def execute(self, name: str, args: dict) -> str:
        args = args or {}
        if name == "declare_stages":
            return self._declare_stages((args or {}).get("stages"))
        p = self._safe_rel(args.get("path", ""))
        if name == "write_file":
            return self._write(p, args)
        if name == "edit_file":
            return self._edit(p, args)
        if name == "delete_file":
            return self._delete(p)
        return f"(unknown tool: {name})"

    def _declare_stages(self, stages) -> str:
        """Validate + stage a `looplab_stages.json` of PRECEDING stages. Reserves the final `score`
        stage for the operator's cmd (appended by the engine), so a Developer can add train/prep work
        but never rewrite the scoring. Returns a clear error string on any problem (nothing staged) so
        the tool loop gets actionable feedback instead of the silent malformed-manifest fallback."""
        import json
        # OPERATOR-declared `cmd.stages` pipelines run VERBATIM: the engine's _resolve_stages takes a
        # valid operator list and never reads the Developer manifest, so "declaring" one here would
        # succeed into a file nobody consumes and the repaired node would re-run the identical
        # pipeline until abandon (P12). Refuse with the real route to a fix.
        if self._operator_stages:
            return ("(refused: this task's pipeline is OPERATOR-declared (`cmd.stages`) and runs "
                    "verbatim; the manifest cannot change it — fix the failing stage's script/code "
                    "instead)")
        # The manifest itself is TOOL-OWNED (validated here, engine-validated again at consume time):
        # gate it on the PROTECT list only — an operator may explicitly protect 'looplab_stages.json'
        # to disable Developer pipelines — NOT on the edit surface. The legacy default surface is
        # ["**/*.py"], which no root .json file can ever match, so the old surface gate made this
        # REQUIRED tool refuse on every legacy repo task (the prompt mandates it for training runs).
        # The surface still governs the STAGE SCRIPTS the manifest points at (write_file), and the
        # declared commands run under the same sandbox tier as the eval — declaring a stage grants
        # nothing an in-surface .py edit (imported by the eval) couldn't already run.
        reason = SurfacePolicy(None, self._protected, self._prefixes,
                               protected_exact=True, check_escapes=False).check("looplab_stages.json")
        if reason is not None:
            return ("(refused: looplab_stages.json is protected — the operator owns the eval; "
                    "you may not declare stages in it)")
        # The shared stage rules (runtime/command_eval.validate_stages) — the SAME validator the
        # engine's _resolve_stages runs at consume time, so a manifest this tool accepts is never
        # silently re-filtered engine-side. The refusal strings stay byte-identical to the original
        # inline loop (the model steers on them): validate_stages returns the bare reason, this site
        # wraps it in its historical "(refused: …)" envelope.
        from looplab.runtime.command_eval import validate_stages
        clean, err = validate_stages(stages, reserved=("score",))
        if err is not None:
            return f"(refused: {err})"
        # The OPERATOR's wall-clock ceiling, held at the one moment correcting it is free. Deliberately
        # NOT inside `validate_stages`: that validator is shared with the OPERATOR's own `cmd.stages`
        # (who owns the budget and may declare what they like) and with `_resolve_stages` at CONSUME
        # time, where a refusal drops the whole manifest — which degrades the node to the score command
        # alone and, on a repo carrying a committed baseline checkpoint, records a number about a model
        # this node never trained. A budget rule must never be able to cause that.
        over_budget = self.stage_budget_refusal(clean)
        if over_budget is not None:
            return over_budget
        miss = _missing_stage_input_paths(clean)      # catch a hallucinated non-existent data path
        if miss:
            return f"(refused: {_missing_paths_feedback(miss)})"
        # F1c, the DECLARATION side: a file already in the working set (a repair's seeded config, or
        # one written earlier this session) that names this manifest's own declared output
        # absolutely, in the source tree. See `manifest_collision_refusal`.
        collision = self.manifest_collision_refusal(clean)
        if collision is not None:
            return collision
        self.files["looplab_stages.json"] = json.dumps({"stages": clean}, indent=1)
        chain = " → ".join(s["name"] for s in clean) + " → score (operator cmd)"
        return f"declared {len(clean)} preceding stage(s): {chain}"

    def _refusal(self, p: str, verb: str):
        """Run the shared SurfacePolicy (tools/patch.py) over an already-canonicalized path and map
        its reason codes onto THIS tool's historical refusal strings (byte-identical — the model
        steers on them). `p` came through `_safe_rel`, which is this site's escape gate — hence
        `check_escapes=False`: _safe_rel's rules differ from patch._escapes (it also strips `./`,
        rejects `~`, and accepts a drive-letter path on POSIX). Protected matching is EXACT and
        case-sensitive here (`protected_exact=True`) — the protect entries arrive pre-normalized
        from RepoTask._protected_names — unlike the diff gate's case-insensitive globs. Prefixes
        pass through VERBATIM (no rstrip); see SurfacePolicy's docstring. Returns None when the
        write may proceed."""
        reason = SurfacePolicy(self._surface, self._protected, self._prefixes,
                               protected_exact=True, check_escapes=False).check(p)
        if reason == SurfacePolicy.PROTECTED:
            # Two distinct situations land in PROTECTED: the operator's eval/scorer files, and a
            # read-only DATA MOUNT (protected defensively so the write refuses VISIBLY — see
            # RepoTask._protected_names). Name the real reason so the model takes the right next
            # step: leave the scorer alone vs write derived data to a different path.
            for nm in self._data_mounts:
                if p == nm or p.startswith(nm + "/"):
                    return (f"(refused: {p} is a read-only data mount; you may not {verb} the "
                            "original — write derived/processed data to a different path)")
            return f"(refused: {p} is protected — the operator owns the eval; you may not {verb} it)"
        if reason is not None:
            return f"(refused: {p} is outside your editable surface: {', '.join(self._surface)})"
        return None

    def _source_root_paths(self, content: str) -> list[str]:
        """Absolute paths in `content` that point INTO an editable repo's SOURCE tree.

        Such a path is wrong by construction and the error is invisible at authoring time. A node
        evaluates in its OWN materialized copy of the editable repo, so a path rooted at the source
        never names anything this node produced — it names the operator's original, whose contents
        belong to whoever last ran something there by hand.

        Measured, on `runs/rubertlite-dr-unified-v6` node 0. The train stage ran correctly and wrote
        its checkpoint exactly where its manifest declared, workdir-relative:
        `vectorsearch/experiments/unified-mnr-t05-b8192-e10_rubert-tiny-lite/final/`. But the
        node's `vectorsearch/configs/config.yaml` set
        `test.retriever.model_settings.checkpoint_path` to
        `/home/jovyan/data/vectorizer-unified/vectorsearch/experiments/<same name>/final` — the
        SOURCE tree, where that directory does not and cannot exist (upstream only ever held an old
        human run's `unified-baseline_rubert-tiny-lite`). So the scorer's `.exists()` check failed,
        as it would on every node and every attempt forever, and the scorer retrained. `overwrite:
        true` then deleted the good checkpoint the train stage had produced. The stage's `expect`
        contract was satisfied and then destroyed by the step that was supposed to measure it.

        Reported as a NOTE on a successful write, never a refusal, and this is the deliberate part.
        The rule "an absolute path into the editable source is wrong" is true for artifacts this
        pipeline PRODUCES and merely questionable for a large untracked input the seed mode does not
        copy — a case with a real, first-class answer (`data:`/`references:` mounts, or
        `seed_mode: "all"`) but not one worth refusing a write over, because a refusal the model
        cannot satisfy costs a whole repair attempt. So it goes back the way the compile error does
        (aider-style: feed it straight back at the moment the model can act), and the fully mechanical
        version of this check — refusing an artifact path that collides with a declared `expect.files`
        entry — is written up separately rather than guessed at here.
        """
        out: list[str] = []
        for _name, root in self._roots:
            r = str(root or "").replace("\\", "/").rstrip("/")
            # A root of "" or "/" would match every absolute path; a relative root cannot appear as
            # an absolute path in generated code at all.
            if len(r) < 2 or not r.startswith("/"):
                continue
            idx = content.find(r + "/")
            if idx >= 0 and r not in out:
                out.append(r)
        return out

    def _collision_refusal(self, content: str, where: str = "") -> Optional[str]:
        """The F1c collision, asked from the WRITE side: does this content name, absolutely and in
        the source tree, a path this node's own staged manifest declares its pipeline writes?

        ORDERING — this is the question docs/29 §F1c said had to be answered before building, and
        the code answers it. `_run` (`adapters/repo_developer.py`) constructs `RepoWriteTools`, then
        for a fresh repo node runs the STAGES phase FIRST, which persists
        `write.files["looplab_stages.json"]` in its `_finalize`; the write/edit tools are composed
        only afterwards, for the PLAN and IMPLEMENT phases. So on every fresh node the manifest is
        already in `self.files` at every `write_file`/`edit_file` call, and DURING the stages phase
        there is nothing to check because that phase's tool set is read-only (`_declare_stages_phase`
        composes `EnvInspectTools` + scouts; the manifest arrives through the phase EMIT, not a write
        tool). For an improve/merge the parent's manifest is pre-loaded at construction, and for a
        REPAIR — which skips the stages phase entirely — `repair_from` pre-loads the FAILING NODE's
        own files, manifest included. The entry's worry that the check "must degrade gracefully
        exactly where it matters most" turns out not to bite: the one case with no manifest at all is
        a node whose stages phase produced nothing AND that has no parent, where there is no
        declaration to contradict and silence is the right answer.

        Verified against the incident: `runs/rubertlite-dr-unified-v6` node 0's `node_created.files`
        is `{looplab_stages.json, vectorsearch/configs/config.yaml, vectorsearch/test.py}` — the
        manifest and the colliding config were staged in the SAME session, manifest first.
        """
        hits = manifest_path_collisions(content, self.files.get("looplab_stages.json", ""),
                                        self._roots)
        return collision_feedback(hits, where=where) if hits else None

    def stage_budget_refusal(self, stages) -> Optional[str]:
        """The operator's WALL-CLOCK ceiling, from the DECLARATION side: does any stage claim a longer
        leash than the whole eval gets? Returns the `(refused: …)`-enveloped message, or None.

        A method on the tools rather than a bare call at each site, for the same reason
        `manifest_collision_refusal` is: the two declaring paths — this tool (implement/repair) and the
        stages phase's `_validate` bounce (authoring) — must apply the identical rule to the identical
        budget, and the phase reaches it through the same `write` object it already holds. The rule
        itself is `command_eval.stage_time_budget_refusal`; nothing is re-derived here, because the
        number a role is TOLD and the number it is HELD TO diverging is the whole defect."""
        from looplab.runtime.command_eval import stage_time_budget_refusal
        msg = stage_time_budget_refusal(stages, self._time_budget)
        return f"(refused: {msg})" if msg is not None else None

    def manifest_collision_refusal(self, stages) -> Optional[str]:
        """The same rule from the DECLARATION side: does a file ALREADY in the working set collide
        with the `expect.files` these stages declare?

        The write-side check alone is order-dependent — it can only see a manifest that is already
        staged — and there are two real orders it misses. A REPAIR seeds the failing node's whole
        working set and re-declares stages without necessarily re-writing the config (the config is
        then never handed to `_write`/`_edit` at all), and within one implement session the model may
        write the config before calling `declare_stages`. Asking the same question when the
        DECLARATION moves closes both, with no second rule: `manifest_path_collisions` is called with
        the arguments swapped round, not reimplemented.

        `looplab_stages.json` itself is skipped as CONTENT — an absolute source path in a stage's own
        argv is the `_missing_stage_input_paths` guard's subject, and reading the manifest as content
        against itself would report a stage command that names its own declared output.
        """
        hits: list[tuple] = []
        text = json.dumps({"stages": list(stages or [])})
        for path, content in sorted(self.files.items()):
            if path == "looplab_stages.json" or not isinstance(content, str):
                continue
            for hit in manifest_path_collisions(content, text, self._roots):
                hits.append(hit)
                # Name the FILE that has to change: from this side the model is holding a manifest,
                # and "some staged file contradicts you" is not something it can act on.
                return (collision_feedback(hits, where=path)
                        .replace("Nothing was staged.", "The stages were NOT declared.")
                        .replace("the content you wrote for", "your already-staged"))
        return None

    def _source_root_note(self, content: str) -> str:
        """The advisory tail appended to a successful write/edit (see `_source_root_paths`)."""
        roots = self._source_root_paths(content)
        if not roots:
            return ""
        return (f" NOTE: this content hard-codes an absolute path inside the editable repo's SOURCE "
                f"tree ({roots[0]}/…). This node runs in its OWN copy of that tree, so such a path "
                "can NEVER name an artifact this node's pipeline produced — it names the operator's "
                "original, which will not contain it, and a 'does it exist' check on it fails on "
                "every node forever. Use a path RELATIVE to the eval workdir for anything your own "
                "stages write or read (the same paths your stage manifest declares in `expect`).")

    @staticmethod
    def _py_syntax_error(path: str, content: str) -> Optional[str]:
        """Auto-validator (aider/Claude-Code style: compile after every edit, feed the error straight
        back). For a *.py result, the first compile() error as "line N: msg", else None. Uses
        compile() (not ast.parse) so it ALSO catches the AST-validation errors ast.parse lets through
        — a repeated keyword arg, `return` outside a function, an unmatched paren, a duplicate param.
        The eval sandbox for a repo task runs on THIS interpreter, so ANY compile error here means the
        code won't run there either — hence ALL of them are hard-rejected, not just indentation (a
        stray `unmatched ')'` crashed a real training run). The rare cost: a Docker tier on a NEWER
        Python could reject valid PEP-695-style syntax — acceptable, and the developer should target
        the run's interpreter anyway."""
        if not path.endswith(".py"):
            return None
        try:
            compile(content, path, "exec")
            return None
        except SyntaxError as e:           # IndentationError/TabError subclass this
            return f"line {e.lineno}: {e.msg}"
        except ValueError as e:            # source with NUL bytes etc. — genuinely unrunnable
            return str(e)[:80]

    def _write(self, p, args: dict) -> str:
        if not p:
            return ("(refused: path must be REPO-RELATIVE and inside the repo — no absolute paths, "
                    "no `..`. Write the eval entrypoint, e.g. write_file path='test_looplab.py'.)")
        refusal = self._refusal(p, "modify")
        if refusal:
            return refusal
        content = args.get("content", "")
        err = self._py_syntax_error(p, content)
        if err is not None:
            return (f"(refused: the content you wrote for {p} is not valid Python — {err}. "
                    "Fix the syntax and write_file again; nothing was staged.)")
        # Same channel and same cost as the syntax refusal above (one tool call, nothing staged),
        # for the same reason: the model can act on it NOW, and the alternative is discovering it
        # after the GPU has been spent. See `_collision_refusal`.
        collision = self._collision_refusal(content, where=p)
        if collision is not None:
            return collision
        self.files[p] = content
        if p in self.deleted:
            self.deleted.remove(p)
        return f"wrote {p} ({len(content)} bytes)" + self._source_root_note(content)

    def _edit(self, p, args: dict) -> str:
        if not p:
            return ("(refused: path must be REPO-RELATIVE and inside the repo — no absolute paths, "
                    "no `..`.)")
        refusal = self._refusal(p, "modify")
        if refusal:
            return refusal
        cur = self._current(p)
        if cur is None:
            return (f"(no such file to edit: {p} — it is neither staged this turn nor in the repo. "
                    "Create it with write_file instead.)")
        search = str(args.get("search") or "")
        replace = str(args.get("replace") or "")
        # Exact-match + whitespace-tolerant line-anchored fallback live in tools/edit_match.py
        # (shared, delicate, test-covered); this method only stages the result.
        new, msg = apply_search_replace(cur, search, replace, path=p)
        if new is None:
            return msg
        # Auto-validate: reject an edit that INTRODUCES a compile error (bad indentation, an unmatched
        # paren, a repeated kwarg — all crashed real runs), but only when the ORIGINAL compiled
        # cleanly, so we never punish the model for editing an already-broken file. The error flies
        # straight back so the model fixes it NOW instead of ~112 min later as a training crash.
        cur_err = self._py_syntax_error(p, cur)          # None => original compiled cleanly
        new_err = self._py_syntax_error(p, new)
        if cur_err is None and new_err is not None:
            return (f"(refused: this edit makes {p} invalid Python — {new_err}. Check the "
                    "indentation/brackets of your `replace` block against the surrounding code and "
                    "try again. Nothing was staged.)")
        # Only the REPLACEMENT text, for the SAME reason the advisory note below reads only
        # `replace`: a colliding line the repo already carried is not this edit's doing, and
        # refusing an unrelated hunk over it is a refusal the model cannot satisfy by editing what
        # it was editing. The declaration-side check (`manifest_collision_refusal`) is what reaches
        # a line that arrived in the seeded working set and is never handed to a write tool.
        collision = self._collision_refusal(replace, where=p)
        if collision is not None:
            return collision
        self.files[p] = new
        if p in self.deleted:
            self.deleted.remove(p)          # an edit resurrects a previously-deleted file
        # Only the REPLACEMENT text, not the whole file: a source-root path the repo already carried
        # is not this edit's doing, and re-reporting it on every unrelated hunk in that file is how a
        # note becomes noise the model learns to skip.
        return msg + self._source_root_note(replace)

    def _delete(self, p) -> str:
        if not p:
            return ("(refused: path must be REPO-RELATIVE and inside the repo — no absolute paths, "
                    "no `..`.)")
        # SAME gates as write_file: a deletion must not remove a protected file (the operator's
        # eval/metric/grader) or reach outside the editable surface. Without these, delete_file
        # was a hole around the write-surface enforcement.
        refusal = self._refusal(p, "delete")
        if refusal:
            return refusal
        self.files.pop(p, None)
        if p not in self.deleted:
            self.deleted.append(p)
        return f"deleted {p}"


def _md_cell(value: object, limit: int) -> str:
    """One markdown table cell: truncated, and with `|` neutralized. EVERY cell goes through this,
    not just the notes column — a single pipe in a spreadsheet label or a non-numeric value shifts
    every column after it, and this content is whatever the operator's xlsx happens to contain."""
    # NEWLINES are collapsed too, not just pipes. openpyxl returns wrap-text / multi-line cell
    # strings verbatim, and an embedded "\n" splits the table ROW mid-line — the same corruption a
    # stray pipe causes, only worse, since the row fragments into separate malformed lines.
    return str(value)[:limit].replace("|", "/").replace("\r\n", " ").replace("\r", " ").replace(
        "\n", " ")


def _xlsx_to_markdown(path: str, *, max_rows: int = 120, cap: int = 9000) -> Optional[str]:
    """Best-effort render of a results spreadsheet to a compact markdown table so an agent can read
    it (an .xlsx is opaque binary otherwise). Rows with numeric cells become table rows; free-text
    rows between them are folded into the preceding row's trailing 'notes' column (that's how these
    experiment logs are usually laid out). Returns None if openpyxl isn't installed or the file can't
    be read — never raises."""
    try:
        import openpyxl  # optional dependency; absent -> skip gracefully
    except Exception:  # noqa: BLE001
        return None
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception:  # noqa: BLE001
        return None

    def _num(x):
        try:
            float(x); return True
        except (TypeError, ValueError):
            return False
    rows = []
    cur = None
    for r in ws.iter_rows(values_only=True):
        c = list(r)
        nums = [x for x in c[1:] if _num(x)]
        if c and c[0] not in (None, "") and nums:                 # a data row (label + numbers)
            cur = {"label": str(c[0]).strip(),
                   "vals": [("" if x is None else (round(float(x), 4) if _num(x) else str(x)))
                            for x in c[1:]],
                   "notes": []}
            rows.append(cur)
        elif cur is not None:                                     # a free-text note -> attach above
            note = " ".join(str(x).strip() for x in c if x not in (None, "")).strip()
            if note:
                cur["notes"].append(note)
        if len(rows) >= max_rows:
            break
    if not rows:
        return None
    ncol = max(len(r["vals"]) for r in rows)
    header = "| label | " + " | ".join(f"c{i+1}" for i in range(ncol)) + " | notes |"
    sep = "|" + "---|" * (ncol + 2)
    lines = [header, sep]
    for r in rows:
        vals = r["vals"] + [""] * (ncol - len(r["vals"]))
        notes = _md_cell("; ".join(r["notes"]), 200)
        lines.append(f"| {_md_cell(r['label'], 60)} | "
                     + " | ".join(_md_cell(v, 200) for v in vals) + f" | {notes} |")
    return "\n".join(lines)[:cap]
