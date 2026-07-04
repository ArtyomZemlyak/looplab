"""RepoTask (kind="repo", ADR-7): the R&D agent works inside an EXISTING repo — it edits
experiment code within an allow-listed surface, and success is measured by running the
OPERATOR'S OWN eval command and reading the metric it emits. The agent never authors the
metric (trust boundary): the eval command + its output files are task-owned and protected
from edits (same mechanism that guards the mlebench grader).

Phase 1: ONE editable repo + read-only references mounted for runtime + an explicit
operator-written `eval_spec`. The agent backend (opencode) edits the repo worktree;
offline / on agent failure a NoOp developer leaves the repo at baseline.

See the architecture study (plans/) for the full workspace/eval model and phases.
"""
from __future__ import annotations

import os
import random
from typing import Optional

from pydantic import BaseModel, Field, field_validator, model_validator

from looplab.core.models import Idea, Node, RunState
from looplab.core.parse import LLMClient
from looplab.agents.roles import LLMResearcher


class ReferenceSpec(BaseModel):
    name: str
    path: str
    mount: bool = False        # copy into the eval workdir (runtime dep) vs context-only


class EditableSpec(BaseModel):
    """One editable repo in a multi-repo workspace (Phase 4). Mounted at `<name>/` under the
    eval workdir; the agent may edit files matching `surface` (globs, relative to the repo)
    and must not overwrite `protect`. The single-repo shorthand `RepoTask.editable_path`
    desugars to one of these mounted at the workspace root (name=".")."""
    name: str                  # subdir under the workspace, e.g. "model" / "data_pipeline"
    path: str                  # the repo to mount + let the agent edit
    surface: list[str] = Field(default_factory=lambda: ["**/*.py"])
    protect: list[str] = Field(default_factory=list)
    # How this repo is materialized into each node's eval workdir (autonomy: Genesis picks from the
    # user's words; "" = fall back to the run-wide Settings.seed_mode). "auto" (the smart default):
    # git-tracked files only when it's a git repo (so a tree bloated with untracked artifacts — model
    # checkpoints, data — is NOT deep-copied per node), else a full copy. "tracked": force git-tracked
    # only. "all": force a full recursive copy (the legacy behavior; use for small repos or when
    # untracked files are needed at eval time).
    seed_mode: str = ""


class EvalSpec(BaseModel):
    """The operator's trusted evaluation (the agent does not author this)."""
    command: list[str]                       # argv, no shell; carries env activation
    cwd: str = "."                           # relative to the node eval workdir
    metric: dict = Field(default_factory=lambda: {"kind": "stdout_json", "key": "metric"})
    params_style: str = "none"               # none | cli_overrides
    timeout: float = 600.0
    # Optional setup command run in the workdir BEFORE the eval each time (e.g.
    # ["pip", "install", "-r", "requirements.txt"]). This is how an e2e Developer's
    # dependency changes take effect reproducibly: the agent edits requirements (in the
    # edit-surface), setup installs them, then the eval runs. Setup failure -> node_failed
    # (stderr fed back to the Developer's repair, so it can fix missing deps).
    setup: list[str] = Field(default_factory=list)
    setup_timeout: float = 600.0
    # Run-level setup: runs ONCE at run start (before the first node), in the editable repo root,
    # NOT per node. Use for a one-time, stable dependency install into the shared interpreter (the
    # autonomy default when deps don't change between experiments) — vs the per-node `setup` above,
    # which reinstalls before EVERY eval (use when the agent edits requirements and each node needs
    # its own deps). Genesis picks which to author from the task's words; both may be set. A failed
    # run_setup aborts the run (env is unusable). Empty -> no run-level setup.
    run_setup: list[str] = Field(default_factory=list)
    run_setup_timeout: float = 1800.0
    # Eval profiles (Phase 2): named override+timeout sets the Researcher selects per node,
    # e.g. {"smoke": {"overrides": ["max_steps=20"], "timeout": 60},
    #       "full":  {"overrides": ["max_steps=2000"], "timeout": 1800}}.
    # Search uses a cheap profile; the confirm phase forces "full".
    profiles: dict[str, dict] = Field(default_factory=dict)
    # Drift cross-check (Phase 4, eval_trust_mode="ratify_freeze_drift"): an INDEPENDENT
    # built-in reader (stdout_json/regex | file_json/regex — never `adapter`) that re-reads
    # the same metric from a source the agent can't forge (e.g. the framework's real stdout).
    # When it can't corroborate the frozen adapter within `drift_tolerance`, the metric is
    # discarded and a `spec_drift` event is recorded. None disables the check.
    cross_check: Optional[dict] = None
    drift_tolerance: float = 1e-6
    # Multi-objective (#5). `metrics`: extra named readers reported alongside the primary
    # (audit/observability), e.g. {"latency_ms": {"kind": "stdout_json", "key": "latency"}}.
    # `constraints`: reader specs with a `max`/`min` bound; a node that violates ANY (or whose
    # constraint value can't be read) is still measured but EXCLUDED from best-selection —
    # "optimize the metric subject to latency_ms <= 100". Operator-owned (trust boundary).
    metrics: dict[str, dict] = Field(default_factory=dict)
    constraints: list[dict] = Field(default_factory=list)

    @field_validator("cross_check")
    @classmethod
    def _cross_check_not_adapter(cls, v):
        from looplab.runtime.command_eval import validate_cross_check
        return validate_cross_check(v)


class RepoResearcher:
    """Minimal proposer for repo tasks: a draft, then improve-from-best. Params are free-
    form (the agent edits code from the rationale); numeric params are optional."""

    def __init__(self, seed: int = 0):
        self.rng = random.Random(seed)

    def propose(self, state: RunState, parent: Optional[Node]) -> Idea:
        if parent is None:
            return Idea(operator="draft", params={},
                        rationale="Establish a working baseline change to the experiment.")
        return Idea(operator="improve", params=dict(parent.idea.params),
                    rationale=(f"Improve on node {parent.id} (metric={parent.metric}). "
                               "Make one focused change to raise the eval metric."))


class RepoParamResearcher:
    """Hyperparameter proposer for the cli_overrides framework mode (Phase 2): random
    within bounds, then Gaussian hill-climb around the best. Tags each Idea with the cheap
    `smoke` eval profile for search (the confirm phase upgrades the leaders to `full`)."""

    def __init__(self, bounds: dict, seed: int = 0, step: float = 0.3):
        self.bounds = bounds
        self.rng = random.Random(seed)
        self.step = step

    def propose(self, state: RunState, parent: Optional[Node]) -> Idea:
        keys = list(self.bounds)
        if parent is None:
            params = {k: round(self.rng.uniform(*self.bounds[k]), 6) for k in keys}
            return Idea(operator="draft", params=params, rationale="random hyperparameters",
                        eval_profile="smoke")
        params = {}
        for k in keys:
            lo, hi = self.bounds[k]
            v = parent.idea.params.get(k, (lo + hi) / 2) + self.rng.gauss(0.0, (hi - lo) * self.step)
            params[k] = round(max(lo, min(hi, v)), 6)
        return Idea(operator="improve", params=params, eval_profile="smoke",
                    rationale=f"perturb node {parent.id} (params={parent.idea.params})")


class NoOpRepoDeveloper:
    """Baseline developer: makes no edits (empty file set). Used offline and as the agent's
    fallback, so a failed/absent agent leaves the repo unmodified and the eval measures the
    baseline rather than poisoning the search."""

    def __init__(self) -> None:
        self.last_files: dict[str, str] = {}

    def implement(self, idea: Idea) -> str:
        self.last_files = {}
        return ""

    def repair(self, idea: Idea, code: str, error: str) -> str:
        return ""


class RepoWriteTools:
    """Write side of the in-house repo developer (the LLM authors/edits files via tools). Writes are
    COLLECTED into `self.files` (path -> content) rather than applied to disk — the orchestrator
    materializes them into the node workdir as the node's files, surface-gated + protected-skipped
    just like an external coding agent's diff. The SAME gates are enforced here so the model gets
    immediate feedback (a refused write) instead of having the edit silently dropped downstream."""

    def __init__(self, surface, protected, prefixes=None, editables=None):
        self.files: dict[str, str] = {}
        self.deleted: list[str] = []
        self._surface = list(surface or [])
        self._protected = set(protected or [])
        self._prefixes = list(prefixes or [])
        # Editable repo roots ({name,path}...) so edit_file can patch a file the node hasn't staged
        # yet: current content = staged overlay first, else the original file on disk.
        self._roots = [(e.get("name") or "", e.get("path")) for e in (editables or []) if e.get("path")]

    def _current(self, p: str):
        """The file's CURRENT content for patching: the staged overlay wins (parent files pre-seeded
        by implement_from, or an earlier write this turn), else the original from an editable root.
        Staged paths are workdir-relative and PREFIXED with the editable's name in multi-editable
        setups (the repo mounts at wd/<name>), so strip the owning prefix before joining its root —
        a bare join would probe <root>/<name>/<file> and read a missing (or wrong) original."""
        if p in self.files:
            return self.files[p]
        from pathlib import Path as _P
        for name, r in self._roots:
            rel = p[len(name) + 1:] if name and name != "." and p.startswith(name + "/") else p
            f = _P(r) / rel
            try:
                if f.is_file():
                    return f.read_text(encoding="utf-8", errors="replace")
            except OSError:
                continue
        return None

    @staticmethod
    def _safe_rel(p: str):
        """Canonicalize to a REPO-RELATIVE path or None. Rejects absolute paths and `..` escapes so
        the agent can only stage files inside the repo it edits — without this, an absolute path like
        `/tmp/x.py` slips past a `**/*.py` surface glob (fnmatch's `*` crosses `/`) and the write is
        silently dropped downstream (it's outside the node workdir)."""
        p = str(p or "").replace("\\", "/").strip()
        while p.startswith("./"):
            p = p[2:]
        if not p or p.startswith("/") or p.startswith("~") or p == ".." \
                or p.startswith("../") or "/../" in p:
            return None
        return p

    def specs(self) -> list[dict]:
        from looplab.tools._base import fn_spec
        return [
            fn_spec("edit_file",
                     "Edit an EXISTING file with a minimal SEARCH/REPLACE patch — STRONGLY PREFERRED "
                     "over write_file for changing existing code: it is far faster and safer than "
                     "re-generating a whole file. `search` must be copied EXACTLY (including "
                     "whitespace/indentation) from the file's current content and must occur exactly "
                     "once; `replace` is its replacement. Make several small edit_file calls for "
                     "several changes. Use write_file only for NEW files.",
                     {"path": {"type": "string", "description": "repo-relative path"},
                      "search": {"type": "string", "description": "exact existing snippet (unique in the file)"},
                      "replace": {"type": "string", "description": "the replacement snippet"}},
                     ["path", "search", "replace"]),
            fn_spec("write_file",
                     "Create or OVERWRITE a file in the experiment repo you are editing. Provide the "
                     "FULL file content (not a diff, not a shell command). Use this ONLY to author the "
                     "eval entrypoint and code edits — NOT to inspect files. Path is REPO-RELATIVE "
                     "(e.g. test_looplab.py); absolute paths and paths outside the repo are rejected.",
                     {"path": {"type": "string", "description": "repo-relative path, e.g. test_looplab.py"},
                      "content": {"type": "string", "description": "the complete file content"}},
                     ["path", "content"]),
            fn_spec("delete_file",
                     "Delete a file you previously wrote in this experiment (within your surface).",
                     {"path": {"type": "string"}}, ["path"]),
        ]

    def execute(self, name: str, args: dict) -> str:
        from looplab.tools.patch import _in_surface
        args = args or {}
        p = self._safe_rel(args.get("path", ""))
        if name == "write_file":
            if not p:
                return ("(refused: path must be REPO-RELATIVE and inside the repo — no absolute paths, "
                        "no `..`. Write the eval entrypoint, e.g. write_file path='test_looplab.py'.)")
            if p in self._protected:
                return f"(refused: {p} is protected — the operator owns the eval; you may not modify it)"
            if not _in_surface(p, self._surface, self._prefixes):
                return f"(refused: {p} is outside your editable surface: {', '.join(self._surface)})"
            content = args.get("content", "")
            self.files[p] = content
            if p in self.deleted:
                self.deleted.remove(p)
            return f"wrote {p} ({len(content)} bytes)"
        if name == "edit_file":
            if not p:
                return ("(refused: path must be REPO-RELATIVE and inside the repo — no absolute paths, "
                        "no `..`.)")
            if p in self._protected:
                return f"(refused: {p} is protected — the operator owns the eval; you may not modify it)"
            if not _in_surface(p, self._surface, self._prefixes):
                return f"(refused: {p} is outside your editable surface: {', '.join(self._surface)})"
            cur = self._current(p)
            if cur is None:
                return (f"(no such file to edit: {p} — it is neither staged this turn nor in the repo. "
                        "Create it with write_file instead.)")
            search = str(args.get("search") or "")
            replace = str(args.get("replace") or "")
            if not search:
                return "(refused: empty `search` — copy the exact snippet you want to replace)"
            n = cur.count(search)
            if n == 0:
                # Whitespace-tolerant fallback: match ignoring trailing spaces on each line (models
                # often lose trailing whitespace when copying a snippet). The match must be LINE-
                # ANCHORED — start at a line boundary and end at end-of-line/EOF — because the
                # replacement swaps WHOLE lines: a mid-line substring hit would silently eat the
                # line's prefix/suffix (verified corruption), and the reported success would hide it.
                def _norm(t: str) -> str:
                    return "\n".join(l.rstrip() for l in t.splitlines())
                cn, sn = _norm(cur), _norm(search)
                idx = cn.find(sn) if sn else -1
                anchored = (idx >= 0 and cn.count(sn) == 1
                            and (idx == 0 or cn[idx - 1] == "\n")
                            and (idx + len(sn) == len(cn) or cn[idx + len(sn)] == "\n"))
                if anchored:
                    pre_lines = cn[:idx].count("\n")
                    s_len = sn.count("\n") + 1          # lines in the MATCHED span (from sn, not search
                    #                                     — a trailing blank line in `search` must not
                    #                                     swallow the next real line)
                    cur_lines = cur.splitlines(keepends=True)
                    tail = "".join(cur_lines[pre_lines + s_len:])
                    last_had_nl = cur_lines[pre_lines + s_len - 1].endswith("\n")
                    rep = replace
                    if rep and last_had_nl and not rep.endswith("\n"):
                        rep += "\n"                     # keep the file's line structure
                    elif rep and not last_had_nl and rep.endswith("\n"):
                        rep = rep[:-1]                  # match-at-EOF without trailing newline: keep it so
                    # empty `replace` = deletion of the matched lines; no stray blank line
                    self.files[p] = "".join(cur_lines[:pre_lines]) + rep + tail
                    if p in self.deleted:
                        self.deleted.remove(p)          # an edit resurrects a previously-deleted file
                    return f"edited {p} (whitespace-tolerant match, 1 hunk applied)"
                first = (search.splitlines() or [""])[0].strip()
                near = next((l for l in cur.splitlines() if first and first in l), "")
                return (f"(no match: `search` was not found in {p} — copy it EXACTLY from the current "
                        f"content{', nearest line: ' + near[:120] if near else ''})")
            if n > 1:
                return f"(ambiguous: `search` occurs {n} times in {p} — include more surrounding lines to make it unique)"
            self.files[p] = cur.replace(search, replace, 1)
            if p in self.deleted:
                self.deleted.remove(p)
            return f"edited {p} (1 hunk applied)"
        if name == "delete_file":
            if not p:
                return ("(refused: path must be REPO-RELATIVE and inside the repo — no absolute paths, "
                        "no `..`.)")
            # SAME gates as write_file: a deletion must not remove a protected file (the operator's
            # eval/metric/grader) or reach outside the editable surface. Without these, delete_file
            # was a hole around the write-surface enforcement.
            if p in self._protected:
                return f"(refused: {p} is protected — the operator owns the eval; you may not delete it)"
            if not _in_surface(p, self._surface, self._prefixes):
                return f"(refused: {p} is outside your editable surface: {', '.join(self._surface)})"
            self.files.pop(p, None)
            if p not in self.deleted:
                self.deleted.append(p)
            return f"deleted {p}"
        return f"(unknown tool: {name})"


def _xlsx_to_markdown(path: str, *, max_rows: int = 120, cap: int = 9000) -> Optional[str]:
    """Best-effort render of a results spreadsheet to a compact markdown table so an agent can read
    it (an .xlsx is opaque binary otherwise). Rows with numeric cells become table rows; free-text
    rows between them are folded into the preceding row's trailing 'notes' column (that's how these
    experiment logs are usually laid out). Returns None if openpyxl isn't installed or the file can't
    be read — never raises."""
    try:
        import openpyxl  # optional dependency; absent -> skip gracefully
    except Exception:  # noqa: BLE001
        return None
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception:  # noqa: BLE001
        return None

    def _num(x):
        try:
            float(x); return True
        except (TypeError, ValueError):
            return False
    rows = []
    cur = None
    for r in ws.iter_rows(values_only=True):
        c = list(r)
        nums = [x for x in c[1:] if _num(x)]
        if c and c[0] not in (None, "") and nums:                 # a data row (label + numbers)
            cur = {"label": str(c[0]).strip(),
                   "vals": [("" if x is None else (round(float(x), 4) if _num(x) else str(x)))
                            for x in c[1:]],
                   "notes": []}
            rows.append(cur)
        elif cur is not None:                                     # a free-text note -> attach above
            note = " ".join(str(x).strip() for x in c if x not in (None, "")).strip()
            if note:
                cur["notes"].append(note)
        if len(rows) >= max_rows:
            break
    if not rows:
        return None
    ncol = max(len(r["vals"]) for r in rows)
    header = "| label | " + " | ".join(f"c{i+1}" for i in range(ncol)) + " | notes |"
    sep = "|" + "---|" * (ncol + 2)
    lines = [header, sep]
    for r in rows:
        vals = r["vals"] + [""] * (ncol - len(r["vals"]))
        notes = ("; ".join(r["notes"])[:200]).replace("|", "/")
        lines.append(f"| {r['label'][:60]} | " + " | ".join(str(v) for v in vals) + f" | {notes} |")
    return "\n".join(lines)[:cap]


class LLMRepoDeveloper:
    """In-house LLM developer for repo tasks — no external coding agent (opencode/aider/…) required.
    It reads the repo with the read-only scout tools and AUTHORS the file(s) the eval needs with
    `write_file`, driven by the shared agentic tool loop. Repo editing was originally an
    external-agent-only path (the in-house repo developer is a NoOp); this lets a repo task run on
    just the in-house LLM. The written files become the node's `last_files`, which the orchestrator
    materializes on top of the seeded tree and evaluates."""

    def __init__(self, client: LLMClient, task, *, parser: str = "tool_call",
                 loop_opts: Optional[dict] = None):
        self.client = client
        self.task = task
        self.parser = parser
        self.loop_opts = dict(loop_opts or {})
        self.brief = task.agent_brief()
        rs = task.repo_spec()
        self._surface = rs["edit_surface"]
        self._protected = rs["protected_names"]
        self._editables = rs["editables"]
        self._prefixes = [e["name"] for e in self._editables if e["name"] not in (".", "")]
        self.last_files: dict[str, str] = {}
        self.last_deleted: list[str] = []

    # Files most useful to PRELOAD verbatim so the agent authors the entrypoint without fumbling with
    # a (truncating) read tool. Order = priority; the rest of the surface is appended within budget.
    _PRELOAD_PRIORITY = ("test.py", "settings.py", "train.py", "to_stf.py", "model.py", "loss.py",
                         "dataset.py", "tokenizing.py", "metrics.py", "inference.py", "README.md")

    def _repo_context(self, per_file: int = 12000, total_budget: int = 90000) -> str:
        """Embed the repo's key source files VERBATIM in the prompt so the agent can author the eval
        entrypoint from them directly — instead of writing throwaway 'cat' scripts to dribble a file
        in through a truncating read tool (the failure mode we hit). Listing first, then prioritized
        full-text files within a char budget."""
        from pathlib import Path as _P
        parts: list[str] = []
        used = 0
        for ed in self._editables:
            root = _P(ed["path"])
            if not root.is_dir():
                continue
            try:
                names = sorted(p.name for p in root.iterdir() if p.is_file())
            except OSError:
                names = []
            parts.append(f"# Repository `{ed['name']}` at {root} — files:\n" + ", ".join(names))
            ordered = [n for n in self._PRELOAD_PRIORITY if n in names] + \
                      [n for n in names if n.endswith((".py", ".yaml", ".yml", ".json"))
                       and n not in self._PRELOAD_PRIORITY]
            for n in ordered:
                if used >= total_budget:
                    break
                fp = root / n
                try:
                    txt = fp.read_text(encoding="utf-8", errors="replace")
                except OSError:
                    continue
                snip = txt[:per_file]
                if len(txt) > per_file:
                    snip += f"\n… (+{len(txt) - per_file} more chars truncated)"
                block = f"\n\n--- {ed['name']}/{n} ---\n{snip}"
                parts.append(block)
                used += len(block)
        return "\n".join(parts)

    def _recipes(self, cap: int = 8000) -> str:
        """Pull the repo's canonical run commands from its README so the agent ORCHESTRATES the
        existing train/convert/test scripts instead of reinventing them (and tripping on the pickled
        dataset's custom classes). Lines that ran a repo `.py` script, captured verbatim with the
        nearest preceding label; the budget keeps the most relevant (earliest) ones."""
        import re
        from pathlib import Path as _P
        rows: list[str] = []
        for ed in self._editables:
            try:
                lines = (_P(ed["path"]) / "README.md").read_text(encoding="utf-8",
                                                                 errors="replace").splitlines()
            except OSError:
                continue
            for i, ln in enumerate(lines):
                s = ln.strip()
                if s.startswith("python ") and re.search(r"\b(train|test|to_stf|tokenizing)\.py\b", s):
                    label = ""
                    for j in range(i - 1, max(i - 4, -1), -1):
                        t = lines[j].strip()
                        if t and not t.startswith("python"):
                            label = t
                            break
                    rows.append((f"# {label}\n" if label else "") + s)
        text, used = [], 0
        for r in rows:
            if used + len(r) > cap:
                break
            text.append(r)
            used += len(r)
        return "\n\n".join(text)

    def _results_context(self, cap: int = 9000) -> str:
        """Surface the repo's PAST-EXPERIMENT / results files so the agent grounds its hyperparameter
        choices in the repo's OWN history (which configs reached which metric) — not just the README.
        Matches files whose name looks like results/experiments/benchmark/scores/leaderboard. Text
        files (.md/.csv/.tsv/.txt) go in verbatim; an .xlsx is rendered to a markdown table best-effort
        (openpyxl optional). De-duped by stem, preferring the text version. Empty when there are none."""
        import re
        from pathlib import Path as _P
        pat = re.compile(r"(result|experiment|benchmark|score|leaderboard)", re.I)
        seen: set[str] = set()
        out: list[str] = []
        used = 0
        for ed in self._editables:
            root = _P(ed["path"])
            if not root.is_dir():
                continue
            try:
                files = sorted((p for p in root.iterdir() if p.is_file() and pat.search(p.name)),
                               key=lambda p: (p.suffix.lower() == ".xlsx", p.name))  # text before xlsx
            except OSError:
                files = []
            for fp in files:
                if used >= cap or fp.stem in seen:
                    continue
                ext = fp.suffix.lower()
                text = None
                if ext in (".md", ".csv", ".tsv", ".txt"):
                    try:
                        text = fp.read_text(encoding="utf-8", errors="replace")
                    except OSError:
                        text = None
                elif ext == ".xlsx":
                    text = _xlsx_to_markdown(str(fp))
                if text:
                    seen.add(fp.stem)
                    snip = text[:max(0, cap - used)]
                    out.append(f"--- {fp.name} ---\n{snip}")
                    used += len(snip)
        return "\n\n".join(out)

    def _emit_spec(self) -> dict:
        from looplab.tools._base import fn_spec
        return fn_spec("done",
                        "Call once the file(s) are written and the eval command would run and print "
                        "its metric. Briefly summarize what you wrote.",
                        {"summary": {"type": "string"}}, [])

    def _run(self, idea: Idea, error: Optional[str] = None,
             base: Optional[dict] = None, base_note: str = "",
             base_deleted: Optional[list] = None) -> str:
        from looplab.agents.agent import drive_tool_loop
        write = RepoWriteTools(self._surface, self._protected, self._prefixes, editables=self._editables)
        if error and (self.last_files or self.last_deleted):   # repair: carry prior state so the
            write.files = dict(self.last_files)                # agent amends it — and so the node's
            write.deleted = list(self.last_deleted)            # recorded deletions aren't lost on
            #   a re-materialization (multi-seed confirm / replay / cross-run import), which would
            #   otherwise resurrect a file the search deleted and measure a different workspace.
        elif base or base_deleted:
            # IMPROVE/REFINE from a parent solution: pre-load the parent's files so untouched ones
            # carry over verbatim (cumulative parent→child diff) — the agent PATCHES, it does not
            # regenerate the whole solution from the pristine repo. Deletions carry too: without
            # them the child's workdir re-seeds the pristine repo with the parent's deleted files
            # RESTORED — a different workspace than "parent + patch".
            write.files = dict(base or {})
            write.deleted = list(base_deleted or [])
        params = ", ".join(f"{k}={v}" for k, v in (idea.params or {}).items()) or "(choose sensible values)"
        from looplab.core.hardware import operational_attention_points
        system = (
            "You improve an existing experiment repository by WRITING code, using ONLY the write_file "
            "tool. You OWN the implementation: the researcher proposed the experiment CONCEPT and "
            "hyperparameters; YOU decide how to realise it in code — which existing scripts to "
            "orchestrate, the stage structure, and how to compute + read the metric. " + self.brief + "\n\n"
            "The repository's current source files are included verbatim below — you have everything "
            "you need; do NOT try to read or inspect files, and do NOT write helper/'cat'/'check' "
            "scripts. To CHANGE an existing file, use edit_file with a minimal SEARCH/REPLACE hunk "
            "(strongly preferred — never re-write a whole existing file). Author the eval entrypoint "
            "the eval command runs (it does not exist yet) by "
            "calling write_file with a REPO-RELATIVE path and the FULL file content. The entrypoint "
            "must run the whole experiment and print the metric as the LAST stdout line (a JSON object "
            "with the required key). ALSO include any related metrics you compute in that SAME JSON "
            "object under their own names (e.g. {\"metric\": <objective>, \"recall@10\": .., \"mrr\": ..}) "
            "— every extra key is recorded and shown alongside the objective; only the required key "
            "drives selection, so report generously. Bake the chosen hyperparameters into the code. Stay within your "
            "editable surface; never write protected or absolute paths. When all files are written and "
            "the eval would succeed, call done.\n\n"
            "For a ROUTINE hyperparameter experiment, prefer ORCHESTRATING the repo's EXISTING scripts "
            "via subprocess (`subprocess.run([sys.executable, 'train.py', ...], check=True)`) and map the "
            "proposed hyperparameters onto the scripts' CLI flags (respect each flag's type — e.g. an int "
            "flag needs an int); custom data formats (e.g. pickled classes) usually only deserialize with "
            "the repo's own loaders, so reuse them. BUT you are NOT limited to that: when the experiment's "
            "idea calls for a STRUCTURAL change — a new loss/objective, an architecture tweak, a data or "
            "feature change, a different training procedure — EDIT the repo's code to make it happen. You "
            "may modify ANY editable file (only the protected files are off-limits); never reject a good "
            "idea just because it needs a code change — implement it. "
            "Use ABSOLUTE paths for inputs that live OUTSIDE the repo (relative `../../...` paths in "
            "the README will not resolve from the eval workdir); mounted inputs appear at ./<name> in "
            "the workdir. When a script already computes + reports the metric (e.g. in a produced "
            "checkpoint filename or a results file), read it from there rather than re-deriving it.\n\n"
            "DEFINITION OF DONE for this node: ONE clean experiment run (exit 0, no errors) that prints "
            "the required metric as the last stdout JSON line. Structure the entrypoint as SEPARATE, "
            "IDEMPOTENT STAGES so a failure in a cheap late stage never repeats an expensive early one:\n"
            "  • TRAIN stage: run the training to a STABLE output path in the workdir (e.g. ./ckpt). At "
            "its start, if a valid checkpoint already exists there, SKIP training and reuse it.\n"
            "  • TEST/METRIC stage: load that checkpoint, evaluate, and print the metric. This stage must "
            "be runnable on its OWN against an existing checkpoint — WITHOUT retraining.\n"
            "The eval is re-run in this SAME workdir after each fix (outputs persist), so when a later "
            "stage (metric parse, conversion, evaluation) fails, your fix + re-run reuses the already-"
            "trained checkpoint and finishes in seconds. NEVER discard a completed training over a "
            "trivial downstream bug, and never silently emit a fake/zero metric to hide an error — fail "
            "loudly (non-zero exit) so the failing stage can be repaired.\n"
            "LOGGING: keep the training framework's logger (e.g. PyTorch Lightning's TensorBoardLogger) "
            "ENABLED and log SEVERAL metrics (the target metric AND related ones — loss, other recalls, "
            "lr), not just the objective; point its log dir at a STABLE path under the workdir so the "
            "curves persist (viewable via `looplab tensorboard <run_dir>`). Also print readable progress "
            "(epoch/step + current metrics) to stdout — it streams to the live eval log.\n\n"
            + operational_attention_points() + "\n\n"
            "=== CANONICAL COMMANDS (from the repo README — adapt paths to absolute + your "
            "hyperparameters) ===\n" + self._recipes() + "\n\n"
            + (("=== PAST EXPERIMENTS / RESULTS (the repo's own history — which configs reached which "
                "metric; use it to pick strong hyperparameters and beat the best) ===\n"
                + _results + "\n\n") if (_results := self._results_context()) else "")
            + "=== REPOSITORY SOURCE ===\n" + self._repo_context())
        user = (f"Experiment concept (the researcher's idea): {idea.rationale}\nHyperparameters to use: {params}.\n"
                "Design and implement the eval entrypoint (and any edits) now with write_file, then call done.")
        if base:
            cap_each, cap_total, used = 8000, 24000, 0
            parts = []
            for name, body in base.items():
                b = str(body or "")[:cap_each]
                if used + len(b) > cap_total:
                    parts.append(f"--- {name} --- (omitted for space)")
                    continue
                used += len(b)
                parts.append(f"--- {name} ---\n{b}")
            user += ("\n\n=== PARENT SOLUTION (your starting point"
                     + (f"; {base_note}" if base_note else "") + ") ===\n"
                     "The files below are this experiment's PARENT — they are already loaded as your "
                     "working set and carry over verbatim unless you change them. AMEND them with "
                     "edit_file (small SEARCH/REPLACE hunks): change ONLY what this idea requires and "
                     "keep everything else as-is. Do NOT rebuild the solution from scratch and do NOT "
                     "re-write whole files that only need a small change.\n\n"
                     + "\n\n".join(parts))
        if error:
            already = ", ".join(self.last_files) or "(none)"
            user += ("\n\nThe PREVIOUS attempt FAILED — fix ONLY the stage that failed (see the error) with "
                     "MINIMAL edit_file hunks on the offending file(s) (re-write a file only if it is beyond patching). This runs in the SAME workdir, so "
                     "any checkpoint/output an earlier stage already produced is STILL THERE: make the "
                     "code reuse it (skip retraining) and go straight to the failing step. Do not start "
                     f"over from scratch. Files you already wrote: {already}.\n"
                     "--- eval error (stderr/stdout tail) ---\n" + error[:4000])
        messages = [{"role": "system", "content": system}, {"role": "user", "content": user}]
        try:
            drive_tool_loop(self.client, write, messages, self._emit_spec(),
                            finalize=lambda a: (a or {}).get("summary", ""),
                            fallback=lambda m: "", **self.loop_opts)
        except Exception as e:  # noqa: BLE001 - never crash the engine on a developer hiccup
            self.last_files = dict(write.files)
            self.last_deleted = list(write.deleted)
            return f"(developer error: {e})"
        self.last_files = dict(write.files)
        self.last_deleted = list(write.deleted)
        return ""

    def implement(self, idea: Idea) -> str:
        return self._run(idea)

    def implement_from(self, idea: Idea, parent) -> str:
        """Improve/refine: start from the PARENT node's solution and patch it (see _run(base=...)).
        Falls back to a from-scratch implement when the parent carries no files AND no deletions
        (e.g. seeded rows)."""
        files = dict(getattr(parent, "files", {}) or {})
        deleted = list(getattr(parent, "deleted", []) or [])
        if not files and not deleted:
            return self._run(idea)
        note = f"parent experiment #{getattr(parent, 'id', '?')}, metric={getattr(parent, 'metric', None)}"
        return self._run(idea, base=files, base_note=note, base_deleted=deleted)

    def repair(self, idea: Idea, code: str, error: str) -> str:
        return self._run(idea, error=error)


class LLMOnboarder:
    """Phase 3 onboarder: the operator gives the framework's command; the Developer writes a
    metric `adapter` (read_metric(workdir)->float) that extracts the metric from whatever
    tracker/logs the run produced (TensorBoard / MLflow / metrics file / stdout). Returns a
    proposal that a human ratifies (then it's frozen + protected). Writing the adapter code
    is the Developer's job — onboarding reuses the same role, not a bespoke agent."""

    _SYS = ("You write a single Python module that reads the FINAL evaluation metric a "
            "training run produced. Output ONLY one ```python``` block defining "
            "`read_metric(workdir: str) -> float`.")

    def __init__(self, client, repo_path, goal, direction, command, timeout):
        self.client = client
        self.repo_path = repo_path
        self.goal = goal
        self.direction = direction
        self.command = command
        self.timeout = timeout

    def _context(self) -> tuple[str, str]:
        """Repo listing + the contents of a few small text files (the entrypoint, configs)
        so the Developer can see the actual metric shape it must read."""
        from pathlib import Path as _P
        root = _P(self.repo_path)
        _skip = {".git", "__pycache__", ".venv", "node_modules", ".mypy_cache", ".pytest_cache"}
        files = [p for p in root.rglob("*")
                 if p.is_file() and _skip.isdisjoint(p.parts)]
        listing = "\n".join(str(p.relative_to(root)) for p in files[:60])
        snippets, exts = [], (".py", ".json", ".yaml", ".yml", ".cfg", ".toml", ".txt")
        for p in files:
            if p.suffix in exts and p.stat().st_size < 4000:
                snippets.append(f"--- {p.relative_to(root)} ---\n"
                                + p.read_text(encoding="utf-8", errors="replace")[:2000])
            if len(snippets) >= 6:
                break
        return listing, "\n\n".join(snippets)

    def __call__(self) -> dict:
        from looplab.core.parse import extract_code
        cmd = " ".join(self.command) or "(the project's training command)"
        listing, snippets = self._context()
        user = (f"Repository files:\n{listing}\n\nKey file contents:\n{snippets}\n\n"
                f"The training command `{cmd}` runs in the work directory. Goal: {self.goal} "
                f"({self.direction}imize). Write `read_metric(workdir)` that, AFTER the run, "
                "returns the final metric by reading whatever the framework wrote (match the "
                "metric key/format you see in the files above — e.g. a JSON like "
                '{"metric": <float>}). Prefer stdlib; if you use an optional tracker lib '
                "(tensorboard/mlflow), import it INSIDE a try/except and fall back. Return a "
                "float; on any problem return a clearly-bad value so the run is not rewarded.")
        try:
            code = extract_code(self.client.complete_text(
                [{"role": "system", "content": self._SYS}, {"role": "user", "content": user}]))
        except Exception as e:  # noqa: BLE001 — propose a stub; human will reject/fix
            code = f"def read_metric(workdir):\n    raise RuntimeError({str(e)!r})\n"
        return {
            "eval_spec": {"command": list(self.command),
                          "metric": {"kind": "adapter", "path": "LOOPLAB_adapter.py"},
                          "params_style": "none", "timeout": self.timeout},
            "adapter_files": {"LOOPLAB_adapter.py": code},
            "goal": self.goal,
        }


class RepoTask(BaseModel):
    kind: str = "repo"
    id: str = "repo_task"
    goal: str = ""
    direction: str = "max"                    # "min" | "max"
    seed: int = 0

    editable_path: str = ""                   # the repo the agent may modify (mounts at root)
    edit_surface: list[str] = Field(default_factory=lambda: ["**/*.py"])
    protect: list[str] = Field(default_factory=list)   # paths the agent must NOT overwrite
    seed_mode: str = ""                        # "" -> Settings.seed_mode | auto | tracked | all
    #  (how the root editable_path is materialized per node; see EditableSpec.seed_mode)
    # Multi-repo workspace (Phase 4): additional editable repos, each mounted at its `name`
    # subdir with its own surface/protect. Use this (optionally with editable_path for a
    # root repo) to let the agent edit across several repos in one experiment.
    editables: list[EditableSpec] = Field(default_factory=list)
    references: list[ReferenceSpec] = Field(default_factory=list)
    data: dict[str, str] = Field(default_factory=dict)  # name -> abs path, copied into eval wd
    eval: Optional[EvalSpec] = None            # operator-given eval; None when onboard=True
    # cli_overrides hyperparameter space (Phase 2): name -> (lo, hi). When set with
    # eval.params_style="cli_overrides", the Researcher tunes these and they become CLI
    # overrides on the eval command — driving an existing framework with NO code edits
    # (developer_backend stays "default" -> the NoOp baseline developer).
    params: dict[str, tuple[float, float]] = Field(default_factory=dict)
    # Onboarding (Phase 3): the agent proposes the trusted eval. The operator gives the
    # framework's command; the Developer writes a metric `adapter` for its tracker; a human
    # ratifies it. `eval` is then left None until ratified.
    onboard: bool = False
    onboard_command: list[str] = Field(default_factory=list)
    onboard_timeout: float = 600.0

    @field_validator("direction")
    @classmethod
    def _direction_valid(cls, v):
        if v not in ("min", "max"):     # silently treating typos as 'minimize' flips the objective
            raise ValueError(f"direction must be 'min' or 'max', got {v!r}")
        return v

    @model_validator(mode="after")
    def _expand_repo_paths(self):
        """Expand ~ and $ENV in every filesystem path the task points at, so a natural
        `editable_path: "~/data/vectorizer/dense-retrieval"` (or "$HOME/…") actually resolves —
        without this the `~` is treated as a literal directory and the repo mounts/scout tools come
        up EMPTY. Runs once; the resolved absolute path is what gets recorded + mounted. (A repo task
        is inherently tied to a local path, so resuming on a DIFFERENT machine/user whose home differs
        already needs editable_path re-pointed — the absolute snapshot just makes that explicit instead
        of silently re-resolving `~` to a different repo.)"""
        exp = lambda p: os.path.expanduser(os.path.expandvars(p)) if isinstance(p, str) and p else p
        self.editable_path = exp(self.editable_path)
        for e in self.editables:
            e.path = exp(e.path)
        for r in self.references:
            r.path = exp(r.path)
        self.data = {k: exp(v) for k, v in self.data.items()}
        return self

    @field_validator("editables")
    @classmethod
    def _names_distinct_and_safe(cls, v):
        seen = set()
        for e in v:
            if e.name in (".", "") or "/" in e.name or "\\" in e.name or ".." in e.name:
                raise ValueError(f"editable name must be a simple subdir, got {e.name!r}")
            if e.name in seen:
                raise ValueError(f"duplicate editable name: {e.name!r}")
            seen.add(e.name)
        return v

    @model_validator(mode="after")
    def _at_least_one_editable(self):
        if not self.editable_path and not self.editables:
            raise ValueError("RepoTask needs an editable source: set `editable_path` "
                             "and/or `editables`.")
        # reference/data mount names are used directly as `wd / name` in _seed_workspace, so
        # they must be safe simple subdir names and not collide with each other or an editable.
        used = {e.name for e in self.editables}
        for label, name in ([("reference", r.name) for r in self.references]
                            + [("data", k) for k in self.data]):
            if not name or "/" in name or "\\" in name or ".." in name.split("/"):
                raise ValueError(f"{label} mount name must be a simple subdir, got {name!r}")
            if name in used:
                raise ValueError(f"mount name collision: {name!r} used by more than one source")
            used.add(name)
        return self

    # ------- TaskAdapter hooks -------
    def assets(self) -> dict[str, str]:
        return {}                              # repo/data are tree-mounted, not flat assets

    def _editable_mounts(self) -> list[dict]:
        """Normalize the single-repo shorthand + the multi `editables` into one list of
        {name, path, surface, protect, seed_mode}. name="." mounts at the workspace root.
        seed_mode is per-repo; "" defers to the run-wide Settings.seed_mode at seed time."""
        out: list[dict] = []
        if self.editable_path:
            out.append({"name": ".", "path": self.editable_path,
                        "surface": list(self.edit_surface), "protect": list(self.protect),
                        "seed_mode": self.seed_mode})
        for e in self.editables:
            out.append({"name": e.name, "path": e.path,
                        "surface": list(e.surface), "protect": list(e.protect),
                        "seed_mode": (e.seed_mode or self.seed_mode)})
        return out

    def eval_spec(self) -> dict:
        return self.eval.model_dump() if self.eval else {}

    def make_onboarder(self, settings):
        """Build the onboarder (Phase 3) when `onboard` is set and a live LLM is available;
        otherwise None (offline runs inject one in tests, or use an explicit eval)."""
        if not self.onboard or settings.backend != "llm":
            return None
        from looplab.adapters.tasks import make_llm_client
        repo_path = self.editable_path or (self.editables[0].path if self.editables else "")
        return LLMOnboarder(make_llm_client(settings), repo_path, self.goal,
                            self.direction, self.onboard_command, self.onboard_timeout)

    @staticmethod
    def _normp(p: str) -> str:
        """Canonicalize a protected path to match git-diff paths (forward slashes, no leading
        './') so exact-string membership in `_write_node_files` actually fires for operator
        entries like './secret.py' or 'src\\secret.py'."""
        p = str(p).replace("\\", "/")
        while p.startswith("./"):
            p = p[2:]
        return p

    def _eval_protected(self) -> list[str]:
        """Files the metric is READ from — the agent must not be able to forge the score by
        writing them. Covers a metric FILE (`file_json`/`file_regex`) AND an agent-authored
        `adapter` module, across EVERY reader the eval uses: the primary `metric`, the extra
        `metrics`, each `constraints` reader, and the drift `cross_check` (review C1 — secondary
        reader paths were left unprotected, so constraint/aux/drift values could be forged).
        Namespaced under the eval `cwd` so the protected name matches the workspace-relative path
        the agent edits."""
        if not self.eval:
            return []
        cwd = self._normp((self.eval.cwd or ".").strip()).strip("/")
        pre = "" if cwd in (".", "") else cwd + "/"
        out: list[str] = []

        def _add(reader) -> None:
            if not isinstance(reader, dict):
                return
            kind = reader.get("kind", "")
            if (kind.startswith("file_") or kind == "adapter") and reader.get("path"):
                out.append(self._normp(pre + reader["path"]))

        _add(self.eval.metric)
        for r in (self.eval.metrics or {}).values():
            _add(r)
        for c in (self.eval.constraints or []):
            _add(c)
        _add(self.eval.cross_check)
        seen: set[str] = set()
        return [p for p in out if not (p in seen or seen.add(p))]   # dedupe, keep order

    def _protected_names(self) -> list[str]:
        """Files the agent's edits may never overwrite, namespaced across all editable repos:
        each repo's `protect` (prefixed by its subdir) + the eval metric file. The onboarding
        adapter is added at ratification time."""
        names: list[str] = []
        for ed in self._editable_mounts():
            pre = "" if ed["name"] in (".", "") else ed["name"].rstrip("/") + "/"
            names += [self._normp(pre + p) for p in ed["protect"]]
        return names + self._eval_protected()

    def agent_brief(self) -> str:
        """Instruction for the editing agent (used by make_roles as the CliAgentDeveloper
        brief): the goal, the editable surface, the protected files, and the eval command.
        Phase 4: when several editable repos are mounted, the surface is namespaced per repo
        subdir so the agent knows where it may edit across the workspace."""
        goal = "maximize" if self.direction == "max" else "minimize"
        surf_globs: list[str] = []
        for ed in self._editable_mounts():
            pre = "" if ed["name"] in (".", "") else ed["name"].rstrip("/") + "/"
            surf_globs += [pre + g for g in ed["surface"]]
        surf = ", ".join(surf_globs)
        prot = ", ".join(self._protected_names()) or "(none)"
        cmd = " ".join(self.eval.command) if self.eval else "(proposed during onboarding)"
        setup = " ".join(self.eval.setup) if (self.eval and self.eval.setup) else ""
        deps = (f" Dependencies are installed before each eval by `{setup}`, so to add a "
                "package, edit the requirements file it reads (if it is in your allowed "
                "paths).") if setup else ""
        return (f"You are improving an existing experiment repository to {goal} the eval "
                f"metric. Goal: {self.goal}\n"
                f"You may ONLY edit files matching: {surf}. Do NOT modify (the operator "
                f"runs the evaluation): {prot}. The eval is run as: `{cmd}`.{deps} Make one "
                f"focused change to make the eval succeed and improve the metric, then stop.")

    def repo_spec(self) -> dict:
        mounts = self._editable_mounts()
        surface: list[str] = []
        for ed in mounts:
            pre = "" if ed["name"] in (".", "") else ed["name"].rstrip("/") + "/"
            surface += [pre + g for g in ed["surface"]]
        return {
            "editables": mounts,                         # Phase 4: every editable repo + mount
            "edit_surface": surface,                     # namespaced union over all repos
            "protected_names": self._protected_names(),
            "references": [r.model_dump() for r in self.references],
            "data": dict(self.data),
            # Back-compat single-seed hint (first editable): consumers that still seed one dir.
            "editable_path": mounts[0]["path"] if mounts else "",
        }

    def _bounds(self) -> dict:
        return {k: (float(lo), float(hi)) for k, (lo, hi) in self.params.items()}

    def build_roles(self):                     # offline: no edits (baseline / param search)
        if self.params:                        # cli_overrides hyperparameter search
            return (RepoParamResearcher(self._bounds(), seed=self.seed), NoOpRepoDeveloper())
        return (RepoResearcher(seed=self.seed), NoOpRepoDeveloper())

    def llm_roles(self, client: LLMClient, parser: str = "tool_call"):
        """When `params` is set: an LLM hyperparameter proposer over the bounds (framework
        cli_overrides mode). Otherwise a free-form code-edit proposer; the editing agent is
        wired in make_roles from repo_spec, and this NoOp developer is the validator's
        baseline fallback."""
        if self.params:
            goal = "maximize" if self.direction == "max" else "minimize"
            hint = (f"Propose hyperparameters to {goal} the eval metric, within bounds: "
                    + ", ".join(f"{k} in [{lo}, {hi}]" for k, (lo, hi) in self._bounds().items()))
            return (LLMResearcher(client, space_hint=hint, bounds=self._bounds(),
                                  parser=parser), NoOpRepoDeveloper())
        hint = ("You are improving an existing experiment repository. Propose the next "
                "concrete change to try (as a short rationale); leave params empty unless "
                "the experiment exposes numeric knobs.")
        return (LLMResearcher(client, space_hint=hint, bounds=None, parser=parser),
                NoOpRepoDeveloper())
