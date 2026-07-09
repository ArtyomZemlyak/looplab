"""The in-house Developer half of the repo task (kind="repo"), split out of
`adapters/repo_task.py` (BACKLOG §4 "repo_task split"): `RepoWriteTools` (the surface-gated
write/edit/delete tool provider whose writes are COLLECTED, not applied), `LLMRepoDeveloper`
(the tool-loop LLM developer that authors/patches the repo's files), `LLMOnboarder` (Phase 3
eval onboarding) and the `_xlsx_to_markdown` results renderer.

A fresh (non-repair) repo implement runs THREE separately-traced phases — STAGES → PLAN →
IMPLEMENT (see `LLMRepoDeveloper._run`): a mandatory READ-ONLY stages phase declares the ordered
eval pipeline (prep → train → … before the operator's protected `score` cmd) via a `declare_stages`
emit and writes `looplab_stages.json`; the plan phase decomposes the code changes into atomic steps;
the implement phase writes the code those stages run. A repair is a single focused session (no
stages/plan). Because stages are owned by this dedicated phase, `declare_stages` is NOT in the
implement write toolset (`RepoWriteTools`) — the manifest is already written before implement starts.

The task/spec half (`RepoTask`, `ReferenceSpec`/`EditableSpec`/`EvalSpec`, the researchers and
`NoOpRepoDeveloper`) stays in `repo_task.py`, which re-imports these names at its END for
back-compat — so `looplab.adapters.repo_task` and the flat `looplab.repo_task` alias keep
exporting them, and this module needs nothing from `repo_task` at import time (no cycle).
"""
from __future__ import annotations

import os
import re
from typing import Optional

from looplab.core.models import Idea
from looplab.core.parse import LLMClient
from looplab.tools.edit_match import apply_search_replace
from looplab.tools.patch import SurfacePolicy

# Absolute paths to INPUT data files referenced in a stage command. Only clear INPUT-data extensions
# (a checkpoint .ckpt/.pt an earlier stage WRITES is deliberately excluded, and relative paths resolve
# to mounts at eval time so are left to the eval). Used to catch the #1 real failure: a train stage
# pointing at a hallucinated argparse-default `.pck` that isn't on this machine.
# The leading `/` must be a TRUE absolute-path boundary — the negative lookbehind rejects a `/` that is
# part of a relative `./dir/...` or `a/b/...` (those resolve to mounts/workdir at eval time, not here).
_INPUT_DATA_RE = re.compile(
    r"(?<![\w./~])/[^\s\"',:]+\.(?:pck|parquet|csv|tsv|npy|npz|pkl|arrow|jsonl|feather|h5|hdf5)")

# Flags whose following token (or `=`-joined value) is an OUTPUT path a stage WRITES — a pipeline
# intermediate that legitimately does not exist yet at declare time (data_prep writes it, train reads it).
_OUTPUT_FLAGS = frozenset({
    "-o", "--out", "--output", "--outdir", "--out-dir", "--out_dir", "--output-dir", "--output_dir",
    "--out-path", "--out_path", "--output-path", "--output_path", "--save", "--save-to", "--save_to",
    "--save-dir", "--save_dir", "--savedir", "--dest", "--destination", "--export", "--write-to"})


def _stage_produced_paths(stages) -> set[str]:
    """ABSOLUTE data paths the pipeline itself WRITES — the token right after an output flag, or a
    `--out=PATH` form. These are intermediates a LATER stage reads, so they must not be flagged as
    "missing input" at declare time just because an EARLIER stage hasn't run yet."""
    produced: set[str] = set()
    for s in (stages or []):
        if not isinstance(s, dict):
            continue
        cmd = s.get("command") or []
        for i, tok in enumerate(cmd):
            if not isinstance(tok, str):
                continue
            val = None
            if tok in _OUTPUT_FLAGS and i + 1 < len(cmd) and isinstance(cmd[i + 1], str):
                val = cmd[i + 1]                                    # `--out PATH`
            elif "=" in tok and tok.split("=", 1)[0] in _OUTPUT_FLAGS:
                val = tok.split("=", 1)[1]                          # `--out=PATH`
            if val:
                produced.update(_INPUT_DATA_RE.findall(val))
    return produced


def _missing_stage_input_paths(stages) -> list[str]:
    """ABSOLUTE input-data paths referenced in stage commands that DON'T exist on disk — almost always
    a hallucinated default (the recurring failure: a train stage's `--train_dataset /…/train.pck` that
    was copied from the repo's argparse default and isn't here). Absolute paths are location-invariant,
    so a declare-time existence check is sound; relative paths (mounts) and `%params%` are skipped.
    Paths the pipeline PRODUCES (an `--out …` of any stage) are excluded — a valid data_prep→train
    pipeline's intermediate legitimately doesn't exist yet at declare time."""
    produced = _stage_produced_paths(stages)
    missing: list[str] = []
    for s in (stages or []):
        if not isinstance(s, dict):
            continue
        for tok in (s.get("command") or []):
            if not isinstance(tok, str) or "%params%" in tok:
                continue
            for m in _INPUT_DATA_RE.findall(tok):
                if m in produced or m in missing:
                    continue
                if not os.path.exists(m):
                    missing.append(m)
    return missing


def _missing_paths_feedback(missing: list[str]) -> str:
    """The actionable bounce message shown to the Developer so it re-declares with a real path."""
    return ("these data paths in your stage command(s) DO NOT EXIST on this machine: "
            + ", ".join(missing[:5]) + ". Do NOT use the repo's DEFAULT argparse dataset paths — they "
            "are the original author's and aren't here. `list_dir` the ACTUAL data (the task's mounted "
            "dataset dir, e.g. ./<mount>, or the absolute dataset path the task/goal gives) and use a "
            "path that EXISTS. (If a path is produced by an EARLIER stage, reference it relatively.)")


class RepoWriteTools:
    """Write side of the in-house repo developer (the LLM authors/edits files via tools). Writes are
    COLLECTED into `self.files` (path -> content) rather than applied to disk — the orchestrator
    materializes them into the node workdir as the node's files, surface-gated + protected-skipped
    just like an external coding agent's diff. The SAME gates are enforced here so the model gets
    immediate feedback (a refused write) instead of having the edit silently dropped downstream."""

    def __init__(self, surface, protected, prefixes=None, editables=None):
        self.files: dict[str, str] = {}
        self.deleted: list[str] = []
        self._surface = list(surface or [])
        self._protected = set(protected or [])
        self._prefixes = list(prefixes or [])
        # Editable repo roots ({name,path}...) so edit_file can patch a file the node hasn't staged
        # yet: current content = staged overlay first, else the original file on disk.
        self._roots = [(e.get("name") or "", e.get("path")) for e in (editables or []) if e.get("path")]

    def _current(self, p: str):
        """The file's CURRENT content for patching: the staged overlay wins (parent files pre-seeded
        by implement_from, or an earlier write this turn), else the original from an editable root.
        Staged paths are workdir-relative and PREFIXED with the editable's name in multi-editable
        setups (the repo mounts at wd/<name>), so strip the owning prefix before joining its root —
        a bare join would probe <root>/<name>/<file> and read a missing (or wrong) original."""
        if p in self.files:
            return self.files[p]
        from pathlib import Path as _P
        for name, r in self._roots:
            rel = p[len(name) + 1:] if name and name != "." and p.startswith(name + "/") else p
            f = _P(r) / rel
            try:
                if f.is_file():
                    return f.read_text(encoding="utf-8", errors="replace")
            except OSError:
                continue
        return None

    @staticmethod
    def _safe_rel(p: str):
        """Canonicalize to a REPO-RELATIVE path or None. Rejects absolute paths and `..` escapes so
        the agent can only stage files inside the repo it edits — without this, an absolute path like
        `/tmp/x.py` slips past a `**/*.py` surface glob (fnmatch's `*` crosses `/`) and the write is
        silently dropped downstream (it's outside the node workdir)."""
        p = str(p or "").replace("\\", "/").strip()
        while p.startswith("./"):
            p = p[2:]
        if not p or p.startswith("/") or p.startswith("~") or p == ".." \
                or p.startswith("../") or "/../" in p:
            return None
        return p

    def specs(self) -> list[dict]:
        from looplab.tools._base import fn_spec
        return [
            fn_spec("edit_file",
                     "Edit an EXISTING file with a minimal SEARCH/REPLACE patch — STRONGLY PREFERRED "
                     "over write_file for changing existing code: it is far faster and safer than "
                     "re-generating a whole file. `search` must be copied EXACTLY (including "
                     "whitespace/indentation) from the file's current content and must occur exactly "
                     "once; `replace` is its replacement. Make several small edit_file calls for "
                     "several changes. Use write_file only for NEW files.",
                     {"path": {"type": "string", "description": "repo-relative path"},
                      "search": {"type": "string", "description": "exact existing snippet (unique in the file)"},
                      "replace": {"type": "string", "description": "the replacement snippet"}},
                     ["path", "search", "replace"]),
            fn_spec("write_file",
                     "Create or OVERWRITE a file in the experiment repo you are editing. Provide the "
                     "FULL file content (not a diff, not a shell command). Use this ONLY to author the "
                     "eval entrypoint and code edits — NOT to inspect files. Path is REPO-RELATIVE "
                     "(e.g. test_looplab.py); absolute paths and paths outside the repo are rejected.",
                     {"path": {"type": "string", "description": "repo-relative path, e.g. test_looplab.py"},
                      "content": {"type": "string", "description": "the complete file content"}},
                     ["path", "content"]),
            fn_spec("delete_file",
                     "Delete a file you previously wrote in this experiment (within your surface).",
                     {"path": {"type": "string"}}, ["path"]),
            # The pipeline is AUTHORED in the Developer's dedicated STAGES phase (its `declare_stages`
            # emit) BEFORE implement — but a write session still needs a validated route to FIX the
            # manifest: a repair whose root cause is a bad stage (wrong argv / too-low timeout) has no
            # other way to change it (write_file refuses under the default *.py surface; without this
            # spec every repair repeats the identical stage failure until abandon — mega-review D1).
            fn_spec("declare_stages",
                     "FIX the eval pipeline manifest (looplab_stages.json). The stages were already "
                     "declared in the STAGES phase — call this ONLY when the failure you are fixing is "
                     "IN the pipeline itself (a stage's command/timeout/name is wrong), passing the "
                     "FULL corrected ordered list of preceding stages (the operator's protected `score` "
                     "step stays appended after them). `%params%` in a command injects this node's "
                     "hyperparameters; give a long `train` a generous `timeout` (seconds). It VALIDATES "
                     "the manifest and reports errors back. Do not use it to re-plan working stages.",
                     {"stages": {"type": "array", "description":
                                 "ordered preceding stages, each {name, command:[argv...], timeout?, check?}"}},
                     ["stages"]),
        ]

    def execute(self, name: str, args: dict) -> str:
        args = args or {}
        if name == "declare_stages":
            return self._declare_stages((args or {}).get("stages"))
        p = self._safe_rel(args.get("path", ""))
        if name == "write_file":
            return self._write(p, args)
        if name == "edit_file":
            return self._edit(p, args)
        if name == "delete_file":
            return self._delete(p)
        return f"(unknown tool: {name})"

    def _declare_stages(self, stages) -> str:
        """Validate + stage a `looplab_stages.json` of PRECEDING stages. Reserves the final `score`
        stage for the operator's cmd (appended by the engine), so a Developer can add train/prep work
        but never rewrite the scoring. Returns a clear error string on any problem (nothing staged) so
        the tool loop gets actionable feedback instead of the silent malformed-manifest fallback."""
        import json
        # The manifest itself is TOOL-OWNED (validated here, engine-validated again at consume time):
        # gate it on the PROTECT list only — an operator may explicitly protect 'looplab_stages.json'
        # to disable Developer pipelines — NOT on the edit surface. The legacy default surface is
        # ["**/*.py"], which no root .json file can ever match, so the old surface gate made this
        # REQUIRED tool refuse on every legacy repo task (the prompt mandates it for training runs).
        # The surface still governs the STAGE SCRIPTS the manifest points at (write_file), and the
        # declared commands run under the same sandbox tier as the eval — declaring a stage grants
        # nothing an in-surface .py edit (imported by the eval) couldn't already run.
        reason = SurfacePolicy(None, self._protected, self._prefixes,
                               protected_exact=True, check_escapes=False).check("looplab_stages.json")
        if reason is not None:
            return ("(refused: looplab_stages.json is protected — the operator owns the eval; "
                    "you may not declare stages in it)")
        # The shared stage rules (runtime/command_eval.validate_stages) — the SAME validator the
        # engine's _resolve_stages runs at consume time, so a manifest this tool accepts is never
        # silently re-filtered engine-side. The refusal strings stay byte-identical to the original
        # inline loop (the model steers on them): validate_stages returns the bare reason, this site
        # wraps it in its historical "(refused: …)" envelope.
        from looplab.runtime.command_eval import validate_stages
        clean, err = validate_stages(stages, reserved=("score",))
        if err is not None:
            return f"(refused: {err})"
        miss = _missing_stage_input_paths(clean)      # catch a hallucinated non-existent data path
        if miss:
            return f"(refused: {_missing_paths_feedback(miss)})"
        self.files["looplab_stages.json"] = json.dumps({"stages": clean}, indent=1)
        chain = " → ".join(s["name"] for s in clean) + " → score (operator cmd)"
        return f"declared {len(clean)} preceding stage(s): {chain}"

    def _refusal(self, p: str, verb: str):
        """Run the shared SurfacePolicy (tools/patch.py) over an already-canonicalized path and map
        its reason codes onto THIS tool's historical refusal strings (byte-identical — the model
        steers on them). `p` came through `_safe_rel`, which is this site's escape gate — hence
        `check_escapes=False`: _safe_rel's rules differ from patch._escapes (it also strips `./`,
        rejects `~`, and accepts a drive-letter path on POSIX). Protected matching is EXACT and
        case-sensitive here (`protected_exact=True`) — the protect entries arrive pre-normalized
        from RepoTask._protected_names — unlike the diff gate's case-insensitive globs. Prefixes
        pass through VERBATIM (no rstrip); see SurfacePolicy's docstring. Returns None when the
        write may proceed."""
        reason = SurfacePolicy(self._surface, self._protected, self._prefixes,
                               protected_exact=True, check_escapes=False).check(p)
        if reason == SurfacePolicy.PROTECTED:
            return f"(refused: {p} is protected — the operator owns the eval; you may not {verb} it)"
        if reason is not None:
            return f"(refused: {p} is outside your editable surface: {', '.join(self._surface)})"
        return None

    @staticmethod
    def _py_syntax_error(path: str, content: str) -> Optional[str]:
        """Auto-validator (aider/Claude-Code style: compile after every edit, feed the error straight
        back). For a *.py result, the first compile() error as "line N: msg", else None. Uses
        compile() (not ast.parse) so it ALSO catches the AST-validation errors ast.parse lets through
        — a repeated keyword arg, `return` outside a function, an unmatched paren, a duplicate param.
        The eval sandbox for a repo task runs on THIS interpreter, so ANY compile error here means the
        code won't run there either — hence ALL of them are hard-rejected, not just indentation (a
        stray `unmatched ')'` crashed a real training run). The rare cost: a Docker tier on a NEWER
        Python could reject valid PEP-695-style syntax — acceptable, and the developer should target
        the run's interpreter anyway."""
        if not path.endswith(".py"):
            return None
        try:
            compile(content, path, "exec")
            return None
        except SyntaxError as e:           # IndentationError/TabError subclass this
            return f"line {e.lineno}: {e.msg}"
        except ValueError as e:            # source with NUL bytes etc. — genuinely unrunnable
            return str(e)[:80]

    def _write(self, p, args: dict) -> str:
        if not p:
            return ("(refused: path must be REPO-RELATIVE and inside the repo — no absolute paths, "
                    "no `..`. Write the eval entrypoint, e.g. write_file path='test_looplab.py'.)")
        refusal = self._refusal(p, "modify")
        if refusal:
            return refusal
        content = args.get("content", "")
        err = self._py_syntax_error(p, content)
        if err is not None:
            return (f"(refused: the content you wrote for {p} is not valid Python — {err}. "
                    "Fix the syntax and write_file again; nothing was staged.)")
        self.files[p] = content
        if p in self.deleted:
            self.deleted.remove(p)
        return f"wrote {p} ({len(content)} bytes)"

    def _edit(self, p, args: dict) -> str:
        if not p:
            return ("(refused: path must be REPO-RELATIVE and inside the repo — no absolute paths, "
                    "no `..`.)")
        refusal = self._refusal(p, "modify")
        if refusal:
            return refusal
        cur = self._current(p)
        if cur is None:
            return (f"(no such file to edit: {p} — it is neither staged this turn nor in the repo. "
                    "Create it with write_file instead.)")
        search = str(args.get("search") or "")
        replace = str(args.get("replace") or "")
        # Exact-match + whitespace-tolerant line-anchored fallback live in tools/edit_match.py
        # (shared, delicate, test-covered); this method only stages the result.
        new, msg = apply_search_replace(cur, search, replace, path=p)
        if new is None:
            return msg
        # Auto-validate: reject an edit that INTRODUCES a compile error (bad indentation, an unmatched
        # paren, a repeated kwarg — all crashed real runs), but only when the ORIGINAL compiled
        # cleanly, so we never punish the model for editing an already-broken file. The error flies
        # straight back so the model fixes it NOW instead of ~112 min later as a training crash.
        cur_err = self._py_syntax_error(p, cur)          # None => original compiled cleanly
        new_err = self._py_syntax_error(p, new)
        if cur_err is None and new_err is not None:
            return (f"(refused: this edit makes {p} invalid Python — {new_err}. Check the "
                    "indentation/brackets of your `replace` block against the surrounding code and "
                    "try again. Nothing was staged.)")
        self.files[p] = new
        if p in self.deleted:
            self.deleted.remove(p)          # an edit resurrects a previously-deleted file
        return msg

    def _delete(self, p) -> str:
        if not p:
            return ("(refused: path must be REPO-RELATIVE and inside the repo — no absolute paths, "
                    "no `..`.)")
        # SAME gates as write_file: a deletion must not remove a protected file (the operator's
        # eval/metric/grader) or reach outside the editable surface. Without these, delete_file
        # was a hole around the write-surface enforcement.
        refusal = self._refusal(p, "delete")
        if refusal:
            return refusal
        self.files.pop(p, None)
        if p not in self.deleted:
            self.deleted.append(p)
        return f"deleted {p}"


def _xlsx_to_markdown(path: str, *, max_rows: int = 120, cap: int = 9000) -> Optional[str]:
    """Best-effort render of a results spreadsheet to a compact markdown table so an agent can read
    it (an .xlsx is opaque binary otherwise). Rows with numeric cells become table rows; free-text
    rows between them are folded into the preceding row's trailing 'notes' column (that's how these
    experiment logs are usually laid out). Returns None if openpyxl isn't installed or the file can't
    be read — never raises."""
    try:
        import openpyxl  # optional dependency; absent -> skip gracefully
    except Exception:  # noqa: BLE001
        return None
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception:  # noqa: BLE001
        return None

    def _num(x):
        try:
            float(x); return True
        except (TypeError, ValueError):
            return False
    rows = []
    cur = None
    for r in ws.iter_rows(values_only=True):
        c = list(r)
        nums = [x for x in c[1:] if _num(x)]
        if c and c[0] not in (None, "") and nums:                 # a data row (label + numbers)
            cur = {"label": str(c[0]).strip(),
                   "vals": [("" if x is None else (round(float(x), 4) if _num(x) else str(x)))
                            for x in c[1:]],
                   "notes": []}
            rows.append(cur)
        elif cur is not None:                                     # a free-text note -> attach above
            note = " ".join(str(x).strip() for x in c if x not in (None, "")).strip()
            if note:
                cur["notes"].append(note)
        if len(rows) >= max_rows:
            break
    if not rows:
        return None
    ncol = max(len(r["vals"]) for r in rows)
    header = "| label | " + " | ".join(f"c{i+1}" for i in range(ncol)) + " | notes |"
    sep = "|" + "---|" * (ncol + 2)
    lines = [header, sep]
    for r in rows:
        vals = r["vals"] + [""] * (ncol - len(r["vals"]))
        notes = ("; ".join(r["notes"])[:200]).replace("|", "/")
        lines.append(f"| {r['label'][:60]} | " + " | ".join(str(v) for v in vals) + f" | {notes} |")
    return "\n".join(lines)[:cap]


# --- LLMRepoDeveloper prompt text, hoisted from the inline literals in `_run` --------------------
# Prompt strings are contracts: every constant below is byte-identical to the original inline
# text — only the seams where runtime values were concatenated (the brief, the attention points,
# recipes/results/source sections, the parent/repair details) became constant boundaries. The
# `{note}`/`{already}` placeholders are `.format`-filled at the exact spots the old f-strings
# interpolated; neither template contains any other brace.
_REPO_DEV_SYSTEM_INTRO = (
    "You improve an existing experiment repository by WRITING code with the write_file and edit_file "
    "tools (edit_file for changes to existing files, write_file for new ones). You OWN the "
    "implementation: the researcher proposed the experiment CONCEPT and "
    "hyperparameters; YOU decide how to realise it in code — which existing scripts to "
    "orchestrate, the stage structure, and how to compute + read the metric. ")
_REPO_DEV_SYSTEM_BODY = (
    "The repository's key source files are PREVIEWED below (each is TRUNCATED to save space). This is "
    "a preview, NOT the full code — to read a whole file or find an exact symbol/flag/signature, use "
    "the read-only repo scouts: read_file(path) for full content (repo-relative, e.g. train.py), "
    "grep(pattern) to find where something is defined across the repo, find_files(root, pattern) / "
    "list_dir(path) to see what exists. Do NOT write helper/'cat'/'check' scripts. "
    "There is NO shell / bash / run-command tool — you CANNOT execute anything yourself: your ONLY "
    "actions are write_file/edit_file (author code) and the read-only scouts below. The eval runs your "
    "code afterwards. (Calling a 'bash'/'run' tool just wastes a turn — it does not exist.) "
    "ALWAYS use REPO-RELATIVE paths for the scouts (e.g. read_file('train.py'), not an absolute "
    "'/home/…/…' path — those are refused). If a grep/read keeps returning the same content, you "
    "already have it: STOP re-reading and act on what you know. "
    "SCOPE: your read/write tools reach ONLY this repo. Data/model files OUTSIDE it (a dataset or "
    "checkpoint mount named in the task) are NOT readable by your tools here — don't try, and don't "
    "hunt for them; just reference their given path in the CODE you write, which CAN open them at "
    "runtime. Need to know the GPUs? call gpu_info (there is no nvidia-smi — you have no shell). "
    "NEVER GUESS a CLI flag / arg name / config key from the truncated preview — grep or "
    "read_file it first (guessing a flag the script doesn't define is the #1 cause of a crash). "
    "Also GROUND every framework API call in the ACTUAL installed environment with the read-only "
    "inspection "
    "tools, instead of guessing (wrong-version APIs are the #1 cause of failed runs): pkg_info(name) "
    "for a package's exact VERSION (e.g. check pytorch-lightning's version before choosing a Trainer "
    "arg — an arg or an accepted value like precision may differ across versions); py_api(dotted) for "
    "a class/function signature or an Enum's VALID VALUES; read_installed(module) to read an installed "
    "module's source; grep_installed(query, package) to find where an arg is parsed / a value "
    "validated. Also: only pass a CLI flag to a repo script if that flag EXISTS in the script's "
    "argparse — CONFIRM it with grep('add_argument') or read_file before you build the "
    "command; otherwise EDIT the script to add it; never invent a flag. "
    "Your write_file/edit_file results are AUTO-VALIDATED (the file is compiled after every change) — "
    "if you get 'not valid Python — line N: …', fix that line immediately; a rejected edit was NOT "
    "staged. To CHANGE an existing file, use edit_file with a minimal SEARCH/REPLACE hunk "
    "(strongly preferred — never re-write a whole existing file). Author the eval entrypoint "
    "the eval command runs (it does not exist yet) by "
    "calling write_file with a REPO-RELATIVE path and the FULL file content. The entrypoint "
    "must print the metric as the LAST stdout line (a JSON object with the required key). CRITICAL: the "
    "eval command runs `<entrypoint>.py`, so THAT FILE MUST EXIST in the workspace after your edits — a "
    "fresh node starts WITHOUT it (unless the operator PROTECTED an existing scorer, which you must NOT "
    "rewrite). For TRAINING work, WHEN the node's declared pipeline (see the task message) has a separate "
    "`train` stage, the entrypoint here only SCORES, and a fixed eval re-runs without "
    "paying to re-train. When NO train stage is declared, the single entrypoint must orchestrate train→test; "
    "editing only train.py leaves the eval with 'no such file: "
    "<entrypoint>.py'. CRITICAL for a TRAINING task: the entrypoint MUST actually TRAIN a model "
    "for THIS experiment (run the repo's train script with your config → produce a FRESH checkpoint) and "
    "THEN score that model. NEVER shortcut by loading a pre-existing/best checkpoint, or by reading a "
    "static results file (a prior run's results_last.csv / *.ckpt is NOT this node's score) — a node that "
    "doesn't train can't test your idea and silently fakes the parent's number. Do NOT add "
    "'skip training if a checkpoint already exists' idempotency: an improve node INHERITS the parent's "
    "checkpoint dir and an interrupted run leaves a partial checkpoint, so skip-if-exists silently reuses "
    "an undertrained / parent model and freezes the metric. Your training code must train UNCONDITIONALLY "
    "when it runs (never self-skip based on a checkpoint file) — reusing a good checkpoint to re-run ONLY "
    "the eval after a fix is the ENGINE's job via the multi-stage pipeline below (separate train/eval "
    "stages), NOT a check inside your script. Ensure the FULL schedule completes (all requested epochs — "
    "the best-val checkpoint of a full run, not an epoch-0/1 checkpoint from a training that never "
    "finished). ALSO include any related "
    "metrics you compute in that SAME JSON "
    "object under their own names (e.g. {\"metric\": <objective>, \"recall@10\": .., \"mrr\": ..}) "
    "— every extra key is recorded and shown alongside the objective; only the required key "
    "drives selection, so report generously. Bake the chosen hyperparameters into the code. Stay within your "
    "editable surface; never write protected or absolute paths. When all files are written and "
    "the eval would succeed, call done.\n\n"
    "TRAIN-THEN-SCORE PIPELINE — the ordered stages are declared in your dedicated STAGES phase and "
    "written to `looplab_stages.json` (the task message states this node's ACTUAL pipeline — trust it, "
    "not an assumption); HERE you implement the CODE those "
    "stages run (e.g. the train.py the `train` stage invokes, the prep.py a `data_prep` stage invokes, the "
    "eval entrypoint the `score` step runs). For reference, a stage is "
    "{name:'train',command:['python','train.py','%params%'],timeout:14400,check:true}; the operator's "
    "`cmd` is APPENDED automatically as the final, protected `score` stage — you CANNOT rewrite how the "
    "run is scored (that's the trust boundary), only add work before it. Stages run in ORDER in the SAME "
    "workdir (artifacts persist: `train` writes a checkpoint the `score` step reads). This is the ONLY "
    "correct way to get 'a failed step is fixed and re-run WITHOUT paying to re-train': the ENGINE reuses "
    "the completed `train` stage's checkpoint and re-runs only what changed (a FRESH node still trains "
    "from scratch — stages are tracked PER NODE, never inherited). Give `train` a GENEROUS `timeout` that "
    "covers the full schedule (epochs × minutes/epoch × 60 — the default is short and would SIGKILL a long "
    "train into an undertrained checkpoint). Put `%params%` inside a stage command to inject THIS node's "
    "hyperparameters as `--key value`, or bake the values into the code yourself. Do NOT hand-roll a "
    "single monolithic entrypoint with a 'skip training if a checkpoint already exists' check: the engine "
    "can't see stage boundaries there, so it can't re-run just the scoring, and that check silently reuses "
    "a PARENT node's / half-finished run's checkpoint (freezing an undertrained model). `declare_stages` "
    "validates your manifest and reports errors back to you. Without stages, your single entrypoint (the "
    "operator's cmd) runs as one command.\n\n"
    "For a ROUTINE hyperparameter experiment, prefer ORCHESTRATING the repo's EXISTING scripts "
    "via subprocess (`subprocess.run([sys.executable, 'train.py', ...], check=True)`) and map the "
    "proposed hyperparameters onto the scripts' CLI flags (respect each flag's type — e.g. an int "
    "flag needs an int); custom data formats (e.g. pickled classes) usually only deserialize with "
    "the repo's own loaders, so reuse them. BUT you are NOT limited to that: when the experiment's "
    "idea calls for a STRUCTURAL change — a new loss/objective, an architecture tweak, a data or "
    "feature change, a different training procedure — EDIT THE REPO'S SOURCE FILES DIRECTLY with "
    "edit_file (e.g. change the loss in train.py/model.py/loss.py with a minimal SEARCH/REPLACE "
    "hunk), then run the training script unchanged. You may modify ANY editable file (only the "
    "protected files are off-limits); never reject a good idea just because it needs a code change "
    "— implement it. "
    "CRITICAL — do NOT make a structural change by generating an entrypoint that REWRITES or "
    "PATCHES another script's source at RUNTIME (string replacement / re.sub / sed / inserting "
    "lines / regex-editing train.py before running it). That pattern reliably corrupts the file "
    "(IndentationError, repeated keyword args, an inserted arg the parser never sees) and the run "
    "fails. Instead make the change PERSISTENT and REVIEWABLE by editing the actual source file "
    "with edit_file, so the training script on disk already contains your change before it runs. "
    "Use ABSOLUTE paths for inputs that live OUTSIDE the repo (relative `../../...` paths in "
    "the README will not resolve from the eval workdir); mounted inputs appear at ./<name> in "
    "the workdir. When a script already computes + reports the metric (e.g. in a produced "
    "checkpoint filename or a results file), read it from there rather than re-deriving it.\n\n"
    "DEFINITION OF DONE for this node: ONE clean experiment run (exit 0, no errors) that prints "
    "the required metric as the last stdout JSON line. Structure the entrypoint as SEPARATE, "
    "IDEMPOTENT STAGES so a failure in a cheap late stage never repeats an expensive early one:\n"
    "  • TRAIN stage: run the training to a STABLE output path in the workdir (e.g. ./ckpt). At "
    "its start, if a valid checkpoint already exists there, SKIP training and reuse it.\n"
    "  • TEST/METRIC stage: load that checkpoint, evaluate, and print the metric. This stage must "
    "be runnable on its OWN against an existing checkpoint — WITHOUT retraining.\n"
    "The eval is re-run in this SAME workdir after each fix (outputs persist), so when a later "
    "stage (metric parse, conversion, evaluation) fails, your fix + re-run reuses the already-"
    "trained checkpoint and finishes in seconds. NEVER discard a completed training over a "
    "trivial downstream bug, and never silently emit a fake/zero metric to hide an error — fail "
    "loudly (non-zero exit) so the failing stage can be repaired.\n"
    "LOGGING: keep the training framework's logger (e.g. PyTorch Lightning's TensorBoardLogger) "
    "ENABLED and log SEVERAL metrics (the target metric AND related ones — loss, other recalls, "
    "lr), not just the objective; point its log dir at a STABLE path under the workdir so the "
    "curves persist (viewable via `looplab tensorboard <run_dir>`). Also print readable progress "
    "(epoch/step + current metrics) to stdout — it streams to the live eval log.\n\n")
_REPO_DEV_COMMANDS_HEADER = (
    "=== CANONICAL COMMANDS (from the repo README — adapt paths to absolute + your "
    "hyperparameters) ===\n")
_REPO_DEV_RESULTS_HEADER = (
    "=== PAST EXPERIMENTS / RESULTS (the repo's own history — which configs reached which "
    "metric; use it to pick strong hyperparameters and beat the best) ===\n")
_REPO_DEV_SOURCE_HEADER = "=== REPOSITORY SOURCE (PREVIEW — truncated; read_file / grep for full) ===\n"
_REPO_DEV_PARENT_BLOCK = (
    "\n\n=== PARENT SOLUTION (your starting point{note}) ===\n"
    "The files below are this experiment's PARENT — they are already loaded as your "
    "working set and carry over verbatim unless you change them. AMEND them with "
    "edit_file (small SEARCH/REPLACE hunks): change ONLY what this idea requires and "
    "keep everything else as-is. Do NOT rebuild the solution from scratch and do NOT "
    "re-write whole files that only need a small change.\n\n")
_REPO_DEV_REPAIR_BLOCK = (
    "\n\nThe PREVIOUS attempt FAILED — fix ONLY the stage that failed (see the error) with "
    "MINIMAL edit_file hunks on the offending file(s) (re-write a file only if it is beyond patching). This runs in the SAME workdir, so "
    "any checkpoint/output an earlier stage already produced is STILL THERE: make the "
    "code reuse it (skip retraining) and go straight to the failing step. Do not start "
    "over from scratch. Files you already wrote: {already}.\n"
    "--- eval error (stderr/stdout tail) ---\n")


class LLMRepoDeveloper:
    """In-house LLM developer for repo tasks — no external coding agent (opencode/aider/…) required.
    It reads the repo with the read-only scout tools and AUTHORS the file(s) the eval needs with
    `write_file`, driven by the shared agentic tool loop. Repo editing was originally an
    external-agent-only path (the in-house repo developer is a NoOp); this lets a repo task run on
    just the in-house LLM. The written files become the node's `last_files`, which the orchestrator
    materializes on top of the seeded tree and evaluates.

    A fresh implement runs THREE separately-traced phases (see `_run`): STAGES (mandatory, first —
    a read-only phase that declares the ordered eval pipeline around the operator's protected `score`
    cmd, writing `looplab_stages.json`), PLAN (read-only atomic-step decomposition), then IMPLEMENT
    (write the code, one bounded session per step). A REPAIR skips both and runs a single session."""

    def __init__(self, client: LLMClient, task, *, parser: str = "tool_call",
                 loop_opts: Optional[dict] = None, plan_decompose: bool = True,
                 plan_min_steps: int = 2, plan_max_steps: int = 8,
                 session_max_turns: int = 500, session_time_budget_s: float = 1200.0):
        self.client = client
        self.task = task
        self.parser = parser
        self.loop_opts = dict(loop_opts or {})
        # C4 plan decomposition + hard per-session backstop (see Settings.developer_*).
        self._plan_decompose = plan_decompose
        self._plan_min_steps = max(2, int(plan_min_steps))
        self._plan_max_steps = max(1, int(plan_max_steps))
        self._session_max_turns = int(session_max_turns)
        self._session_time_budget_s = float(session_time_budget_s)
        self.brief = task.agent_brief()
        rs = task.repo_spec()
        self._surface = rs["edit_surface"]
        self._protected = rs["protected_names"]
        self._editables = rs["editables"]
        self._prefixes = [e["name"] for e in self._editables if e["name"] not in (".", "")]
        self.last_files: dict[str, str] = {}
        self.last_deleted: list[str] = []

    # Files most useful to PRELOAD verbatim so the agent authors the entrypoint without fumbling with
    # a (truncating) read tool. Order = priority; the rest of the surface is appended within budget.
    # PROVENANCE / HEURISTIC ONLY: these names (incl. the repo-specific `to_stf.py`/`tokenizing.py`)
    # come from the reference repo LoopLab was first exercised on. They are a soft *ordering* prior,
    # not a requirement — an absent name simply doesn't preload, and the full surface is appended
    # anyway — so the heuristic degrades gracefully on any other repo. Generalize to an
    # `EditableSpec.preload_priority` knob if a task ever needs to override the order.
    _PRELOAD_PRIORITY = ("test.py", "settings.py", "train.py", "to_stf.py", "model.py", "loss.py",
                         "dataset.py", "tokenizing.py", "metrics.py", "inference.py", "README.md")

    def _repo_context(self, per_file: int = 3000, total_budget: int = 30000) -> str:
        """Embed the repo's key source files VERBATIM in the prompt so the agent can author the eval
        entrypoint from them directly — instead of writing throwaway 'cat' scripts to dribble a file
        in through a truncating read tool (the failure mode we hit). Listing first, then prioritized
        full-text files within a char budget."""
        from pathlib import Path as _P
        parts: list[str] = []
        used = 0
        for ed in self._editables:
            root = _P(ed["path"])
            if not root.is_dir():
                continue
            try:
                names = sorted(p.name for p in root.iterdir() if p.is_file())
            except OSError:
                names = []
            parts.append(f"# Repository `{ed['name']}` at {root} — files:\n" + ", ".join(names))
            ordered = [n for n in self._PRELOAD_PRIORITY if n in names] + \
                      [n for n in names if n.endswith((".py", ".yaml", ".yml", ".json"))
                       and n not in self._PRELOAD_PRIORITY]
            for n in ordered:
                if used >= total_budget:
                    break
                fp = root / n
                try:
                    txt = fp.read_text(encoding="utf-8", errors="replace")
                except OSError:
                    continue
                snip = txt[:per_file]
                if len(txt) > per_file:
                    snip += f"\n… (+{len(txt) - per_file} more chars truncated)"
                block = f"\n\n--- {ed['name']}/{n} ---\n{snip}"
                parts.append(block)
                used += len(block)
        return "\n".join(parts)

    def _recipes(self, cap: int = 8000) -> str:
        """Pull the repo's canonical run commands from its README so the agent ORCHESTRATES the
        existing train/convert/test scripts instead of reinventing them (and tripping on the pickled
        dataset's custom classes). Lines that ran a repo `.py` script, captured verbatim with the
        nearest preceding label; the budget keeps the most relevant (earliest) ones."""
        import re
        from pathlib import Path as _P
        rows: list[str] = []
        for ed in self._editables:
            try:
                lines = (_P(ed["path"]) / "README.md").read_text(encoding="utf-8",
                                                                 errors="replace").splitlines()
            except OSError:
                continue
            for i, ln in enumerate(lines):
                s = ln.strip()
                # The script-name allow-list is a HEURISTIC (train/test are generic; `to_stf`/
                # `tokenizing` are from the first reference repo — see `_PRELOAD_PRIORITY`). It only
                # decides which README command lines get surfaced as recipes; a repo without these
                # names just yields no recipes here, no failure. Widen the pattern if a new repo's
                # entrypoints are missed.
                if s.startswith("python ") and re.search(r"\b(train|test|to_stf|tokenizing)\.py\b", s):
                    label = ""
                    for j in range(i - 1, max(i - 4, -1), -1):
                        t = lines[j].strip()
                        if t and not t.startswith("python"):
                            label = t
                            break
                    rows.append((f"# {label}\n" if label else "") + s)
        text, used = [], 0
        for r in rows:
            if used + len(r) > cap:
                break
            text.append(r)
            used += len(r)
        return "\n\n".join(text)

    def _results_context(self, cap: int = 9000) -> str:
        """Surface the repo's PAST-EXPERIMENT / results files so the agent grounds its hyperparameter
        choices in the repo's OWN history (which configs reached which metric) — not just the README.
        Matches files whose name looks like results/experiments/benchmark/scores/leaderboard. Text
        files (.md/.csv/.tsv/.txt) go in verbatim; an .xlsx is rendered to a markdown table best-effort
        (openpyxl optional). De-duped by stem, preferring the text version. Empty when there are none."""
        import re
        from pathlib import Path as _P
        pat = re.compile(r"(result|experiment|benchmark|score|leaderboard)", re.I)
        seen: set[str] = set()
        out: list[str] = []
        used = 0
        for ed in self._editables:
            root = _P(ed["path"])
            if not root.is_dir():
                continue
            try:
                files = sorted((p for p in root.iterdir() if p.is_file() and pat.search(p.name)),
                               key=lambda p: (p.suffix.lower() == ".xlsx", p.name))  # text before xlsx
            except OSError:
                files = []
            for fp in files:
                if used >= cap or fp.stem in seen:
                    continue
                ext = fp.suffix.lower()
                text = None
                if ext in (".md", ".csv", ".tsv", ".txt"):
                    try:
                        text = fp.read_text(encoding="utf-8", errors="replace")
                    except OSError:
                        text = None
                elif ext == ".xlsx":
                    text = _xlsx_to_markdown(str(fp))
                if text:
                    seen.add(fp.stem)
                    snip = text[:max(0, cap - used)]
                    out.append(f"--- {fp.name} ---\n{snip}")
                    used += len(snip)
        return "\n\n".join(out)

    def _emit_spec(self) -> dict:
        from looplab.tools._base import fn_spec
        return fn_spec("done",
                        "Call once the file(s) are written and the eval command would run and print "
                        "its metric. Briefly summarize what you wrote.",
                        {"summary": {"type": "string"}}, [])

    def _session_opts(self, *, max_turns=None, time_budget=None) -> dict:
        """loop_opts + the HARD per-session ceiling. A developer session ALWAYS gets a finite bound so
        a model that keeps writing/exploring without ever emitting `done` fails cleanly with the code
        it has written, instead of the 10k-call / multi-hour runaway a big task produced."""
        opts = dict(getattr(self, "loop_opts", {}) or {})
        opts["max_turns"] = int(max_turns if max_turns is not None
                                else getattr(self, "_session_max_turns", 500))
        opts["time_budget_s"] = float(time_budget if time_budget is not None
                                      else getattr(self, "_session_time_budget_s", 1200.0))
        return opts

    def _plan_emit_spec(self) -> dict:
        from looplab.tools._base import fn_spec
        return fn_spec("propose_plan",
                        "Propose an ORDERED plan of ATOMIC implementation steps for this experiment. "
                        "Each step is ONE self-contained, independently-verifiable change (e.g. 'add the "
                        "second-stage fine-tune loop to train.py', 'wire the stage-2 hyperparameters', "
                        "'write the eval entrypoint that prints the metric'). Prefer 2-6 SMALL steps; use "
                        "a single step only if the change is genuinely trivial. Do NOT write code here — "
                        "plan only. Call this exactly once when the plan is ready.",
                        {"steps": {"type": "array", "items": {"type": "object", "properties": {
                            "title": {"type": "string", "description": "short imperative title"},
                            "detail": {"type": "string", "description": "concretely what to change and why"}},
                            "required": ["title"]}}},
                        ["steps"])

    def _propose_plan(self, system: str, idea: Idea, write=None) -> list:
        """Plan phase: a READ-ONLY stage — the developer inspects the real code/experiments (it CANNOT
        write here), and its only exit is `propose_plan` (the ordered atomic plan). Returns a list of
        {title, detail}; [] on empty/failure so the caller falls back to one session."""
        from looplab.agents.agent import run_phase, CompositeTools
        from looplab.tools.env_inspect import EnvInspectTools
        params = ", ".join(f"{k}={v}" for k, v in (idea.params or {}).items()) or "(choose sensible values)"
        plan_user = (
            f"Experiment concept (the researcher's idea): {idea.rationale}\nHyperparameters: {params}.\n"
            "This is the PLANNING stage. You can READ and inspect the repo (read_file — it paginates, so "
            "read a file ONCE, don't re-read; grep, find_files, list_dir, pkg_info, py_api, gpu_info) but "
            "you CANNOT write code yet. Actually READ the relevant source (the eval/entry script, the "
            "files you'll change) and any prior experiment you're building on — enough to know EXACTLY "
            "what to change — THEN call propose_plan with an ordered list of ATOMIC, independently-"
            "testable steps, each naming concretely what to change and why. Do NOT guess from the "
            "truncated preview; the implement stage (and update_plan) come next.")
        messages = [{"role": "system", "content": system}, {"role": "user", "content": plan_user}]
        # READ-ONLY toolset: repo scouts + env inspection, but NO write tools — the plan stage's only
        # output is the plan. (This used to be tools=None to force convergence, which made the planner
        # work BLIND off the truncated preview; the read_file pagination fix + emit_after/emit_force
        # convergence backstop now let it read PROPERLY without exploring forever.)
        read_only = CompositeTools([EnvInspectTools()] + self._scout_tools(write))
        try:
            # Full session budget — same contract as every other phase: the soft nudge at
            # agent_emit_after (300) and the forced emit at agent_emit_force (500) ride in via
            # loop_opts, and budget exhaustion salvages a forced emit. The old tight clamp
            # (40 turns / 360s) starved the planner on a big repo the same way it starved the
            # stages phase (read the repo for the whole budget, degrade to []).
            plan = run_phase(
                self.client, read_only, messages, self._plan_emit_spec(),
                label="Developer·plan", next_label="the implement phase",
                finalize=lambda a: (a or {}).get("steps", []), fallback=lambda m: [],
                **self._session_opts())
        except Exception:  # noqa: BLE001 — a failed plan phase just degrades to a single session
            return []
        steps = []
        for s in (plan or [])[: getattr(self, "_plan_max_steps", 8)]:
            if isinstance(s, dict) and (s.get("title") or s.get("detail")):
                steps.append({"title": str(s.get("title", "")).strip(),
                              "detail": str(s.get("detail", "")).strip()})
        return steps

    def _run_step(self, idea: Idea, step: dict, idx: int, total: int, write, system: str,
                  stage_note: str = "") -> str:
        """Execute ONE atomic plan step in a FRESH bounded session, on top of the files accumulated so
        far (carried in `write.files`; syntax is validated per write by the write tool). A step's own
        error never aborts the plan — later steps + the eval still run on whatever got written.
        `stage_note` restates the node's ACTUAL declared pipeline (or its absence) so a step session
        never assumes a train stage the stages phase didn't produce."""
        from looplab.agents.agent import run_phase, CompositeTools
        from looplab.tools.env_inspect import EnvInspectTools
        done_so_far = ", ".join(write.files) or "(none yet)"
        step_user = (
            f"You are implementing a multi-step plan — STEP {idx} of {total}.\n"
            f"Overall experiment: {idea.rationale}\n{stage_note}\n"
            f"THIS STEP — {step['title']}:\n{step.get('detail') or step['title']}\n\n"
            f"Files CURRENTLY in the workspace (the parent solution + whatever earlier steps wrote — read "
            f"any of them with read_file to see their real content, do NOT assume): {done_so_far}\n"
            "Make ONLY the edits THIS step needs with write_file/edit_file — PATCH existing files, don't "
            "regenerate untouched ones — then call done. Do the minimum for this step; later steps handle "
            "the rest. If this is the last step, make sure the eval entrypoint runs end-to-end.")
        messages = [{"role": "system", "content": system}, {"role": "user", "content": step_user}]
        try:
            # implement steps CONSUME the stages/plan briefs + share the node read-cache, but don't
            # contribute (their writes add length faster than signal, and the last step is terminal) —
            # so the ledger stays the 3 exploration briefs (propose/stages/plan), never K-step bloat.
            run_phase(self.client, CompositeTools([write, EnvInspectTools()] + self._scout_tools(write)),
                      messages, self._emit_spec(), label=f"Developer·implement step {idx}/{total}",
                      handoff=False, finalize=lambda a: (a or {}).get("summary", ""),
                      fallback=lambda m: "", **self._session_opts())
        except Exception as e:  # noqa: BLE001
            return f"(step {idx} error: {e})"
        return ""

    def _scout_tools(self, write=None):
        """Read-only repo scouts (read_file / grep / find_files / list_dir) so the Developer can READ
        the code it is EDITING and VERIFY an exact CLI flag / function signature / config key in the
        ACTUAL source instead of GUESSING it — guessing an arg the embedded (truncated) source didn't
        show is a top cause of a training crash. Reuses the SHARED RepoScoutTools (path-safe +
        secret-filtered), bound to the editable repo roots with repo-relative paths (the SAME paths as
        write_file/edit_file). `write.files` is passed as the STAGED overlay so read/grep see the code
        the Developer is currently writing — not the pristine on-disk repo (reading a parent/merge
        source is a separate, secondary concern)."""
        roots = [e["path"] for e in (getattr(self, "_editables", None) or []) if e.get("path")]
        if not roots:
            return []
        from looplab.tools.reposcout import RepoScoutTools
        overlay = write.files if write is not None else None      # live dict the write tools mutate
        deleted = write.deleted if write is not None else None    # staged deletions hidden from read/grep/list
        # (name, path) per editable — MIRRORS RepoWriteTools._roots so a scout hit is rendered/deduped with
        # the SAME `<name>/rel` key the write tools use in a multi-editable repo (round-trips into an edit).
        named = [(e.get("name") or "", e["path"]) for e in (getattr(self, "_editables", None) or []) if e.get("path")]
        return [RepoScoutTools(roots=roots, default_root=roots[0], overlay=overlay, deleted=deleted,
                               named_roots=named)]

    def _stages_emit_spec(self) -> dict:
        from looplab.tools._base import fn_spec
        return fn_spec("declare_stages",
                        "Declare the ORDERED pipeline stages for this experiment and finish the stages "
                        "phase. Each stage is {name, command:[argv...], timeout?, check?}; they run IN "
                        "ORDER in the same workdir so artifacts (a trained checkpoint, prepared data) "
                        "persist to later stages. Put `%params%` in a command to inject THIS node's "
                        "hyperparameters as `--key value`, or bake the values into the argv yourself. "
                        "Give a long training stage a GENEROUS timeout (seconds).",
                        {"stages": {"type": "array", "items": {"type": "object", "properties": {
                            "name": {"type": "string"},
                            "command": {"type": "array", "items": {"type": "string"}},
                            "timeout": {"type": "number"}, "check": {"type": "boolean"}},
                            "required": ["name", "command"]}}},
                        ["stages"])

    def _cmd_context(self) -> tuple[dict, bool]:
        """The operator's scoring contract (eval_spec) + whether one exists. The stages phase shows it to
        the Developer as IMMUTABLE (the engine appends it as the final protected `score` stage); with no
        cmd the Developer must declare the FULL pipeline including a final scoring stage."""
        ev = {}
        try:
            ev = self.task.eval_spec() or {}
        except Exception:  # noqa: BLE001 — a task without eval_spec (toy/tests) => no cmd, full pipeline
            ev = {}
        # Onboard mode: `eval` is None until the adapter is ratified, but the onboard COMMAND is the scorer
        # (the frozen metric adapter reads ITS output). Treat it as the immutable cmd so the stages phase
        # declares PRECEDING train/prep stages around it — NOT a full pipeline whose own score stage would
        # fight the onboarder's adapter (that broke the onboarding run: finished=False).
        if not ev.get("command") and not ev.get("stages"):
            oc = getattr(self.task, "onboard_command", None)
            if oc:
                ev = {**ev, "command": list(oc)}
        has_cmd = bool(ev.get("command") or ev.get("stages"))
        return ev, has_cmd

    def _stages_user(self, idea: Idea, ev: dict, has_cmd: bool) -> str:
        import json as _json
        params = ", ".join(f"{k}={v}" for k, v in (idea.params or {}).items()) or "(bake sensible values)"
        if has_cmd:
            cmd_desc = _json.dumps(ev.get("stages") or ev.get("command"), ensure_ascii=False)[:800]
            metric = _json.dumps(ev.get("metric"), ensure_ascii=False)[:200]
            contract = (
                f"The operator's SCORING command is FIXED (you may NOT change it): `{cmd_desc}`; it reads "
                f"the metric via {metric}. The engine appends it as the final, protected `score` stage. "
                "Your job: declare the ordered stages that run BEFORE it (do NOT include a `score` stage — "
                "it's reserved), producing whatever that scorer reads (a trained checkpoint, prepared data).")
        else:
            contract = (
                "There is NO operator scoring command — declare the FULL pipeline, INCLUDING a final stage "
                "that runs the evaluation and PRINTS the metric the task's metric reader parses. Name that "
                "final stage e.g. `evaluate` — the name `score` is RESERVED (it always denotes an "
                "engine-appended operator step and will be rejected).")
        return (
            f"Experiment concept (the researcher's idea): {idea.rationale}\nHyperparameters for THIS node: "
            f"{params}.\n\nThis is the STAGES phase (first). {contract}\n\n"
            "READ the repo to ground the stages in the ACTUAL entry scripts/args (read_file paginates — "
            "read a file ONCE; grep, find_files, list_dir, pkg_info, py_api). GOOD PRACTICE: separate "
            "stages for data/feature PREPARATION, TRAINING (fresh model every node — never reuse a "
            "checkpoint), and TESTING; bake this node's hyperparameters into the `train` command (or use "
            "`%params%`). Give training a generous timeout. Then call `declare_stages` once. You are NOT "
            "writing code yet — the plan + implement phases come next.")

    def _declare_stages_phase(self, idea: Idea, write, system: str) -> list:
        """Stages phase (MANDATORY, FIRST): a READ-ONLY phase where the Developer studies the repo + the
        operator's cmd and emits `declare_stages` — the ordered pipeline (prep → train → …) that runs
        before the protected `score` step. Writes `looplab_stages.json`. Returns the clean stage list ([]
        on failure — the eval then falls back to just the operator cmd)."""
        from looplab.agents.agent import run_phase, CompositeTools
        from looplab.tools.env_inspect import EnvInspectTools
        from looplab.runtime.command_eval import validate_stages
        import json as _json
        ev, has_cmd = self._cmd_context()
        reserved = ("score",)   # `score` is ALWAYS the engine-appended final stage — consume-side reserves it too
        messages = [{"role": "system", "content": system},
                    {"role": "user", "content": self._stages_user(idea, ev, has_cmd)}]
        # scouts read the LIVE overlay (the parent solution on improve/merge), not the pristine repo
        read_only = CompositeTools([EnvInspectTools()] + self._scout_tools(write))

        def _validate(args):                      # bounce a malformed manifest back to the model
            stages = (args or {}).get("stages")
            _, err = validate_stages(stages, reserved=reserved)
            if err:
                return err
            miss = _missing_stage_input_paths(stages)   # a hallucinated non-existent data path → re-declare
            return _missing_paths_feedback(miss) if miss else None

        def _finalize(args):
            clean, _ = validate_stages((args or {}).get("stages"), reserved=reserved)
            # Mirror `_validate`'s missing-path guard: the force-emit / exhaustion paths call finalize
            # WITHOUT re-running validate, so without this a manifest referencing a hallucinated,
            # non-produced input path would be PERSISTED after the model's retries were exhausted and
            # ship a FileNotFoundError pipeline. Degrade to the operator cmd (return []) instead.
            if clean and not _missing_stage_input_paths(clean):
                write.files["looplab_stages.json"] = _json.dumps({"stages": clean}, indent=1)
                return clean
            return []
        try:
            # Full session budget — the old tight clamp (30 turns / 300s) starved this phase on a
            # big repo: it read for the whole budget, never reached declare_stages, and silently
            # degraded to "no stages declared" (the node then evaluated as a bare single command —
            # observed live). The soft nudge (agent_emit_after=300) / forced emit (agent_emit_force
            # =500) convergence backstop + exhaustion salvage now bound it like every other phase.
            return run_phase(
                self.client, read_only, messages, self._stages_emit_spec(),
                label="Developer·stages", next_label="the plan & implement phases",
                finalize=_finalize, fallback=lambda m: [], validate=_validate,
                **self._session_opts()) or []
        except Exception:  # noqa: BLE001 — a failed stages phase degrades to the operator cmd alone
            return []

    def _run(self, idea: Idea, error: Optional[str] = None,
             base: Optional[dict] = None, base_note: str = "",
             base_deleted: Optional[list] = None) -> str:
        from looplab.agents.agent import run_phase
        from looplab.core import tracing
        write = RepoWriteTools(self._surface, self._protected, self._prefixes, editables=self._editables)
        if base is not None or base_deleted is not None:
            # An EXPLICIT base is the node's OWN solution — the parent's (improve/refine via
            # implement_from) or the failing node's (repair via repair_from). Pre-load it so untouched
            # files carry over verbatim (cumulative diff — the agent PATCHES, doesn't regenerate from
            # the pristine repo) and deletions carry too (else the workdir re-seeds the pristine repo
            # with a deleted file RESTORED). This WINS over `last_files` even for a repair, because the
            # shared developer instance's `last_files` holds whatever node it BUILT LAST — almost never
            # the node being repaired (the create-batch builds every node before any eval).
            write.files = dict(base or {})
            write.deleted = list(base_deleted or [])
        elif error and (self.last_files or self.last_deleted):   # legacy repair (no explicit base):
            write.files = dict(self.last_files)                  # best-effort carry of the last build
            write.deleted = list(self.last_deleted)
        params = ", ".join(f"{k}={v}" for k, v in (idea.params or {}).items()) or "(choose sensible values)"
        from looplab.core.hardware import operational_attention_points
        system = (
            _REPO_DEV_SYSTEM_INTRO + self.brief + "\n\n"
            + _REPO_DEV_SYSTEM_BODY
            + operational_attention_points() + "\n\n"
            + _REPO_DEV_COMMANDS_HEADER + self._recipes() + "\n\n"
            + ((_REPO_DEV_RESULTS_HEADER + _results + "\n\n")
               if (_results := self._results_context()) else "")
            + _REPO_DEV_SOURCE_HEADER + self._repo_context())
        user = (f"Experiment concept (the researcher's idea): {idea.rationale}\nHyperparameters to use: {params}.\n"
                "Design and implement the eval entrypoint (and any edits) now with write_file, then call done.")
        if base:
            cap_each, cap_total, used = 8000, 24000, 0
            parts = []
            for name, body in base.items():
                b = str(body or "")[:cap_each]
                if used + len(b) > cap_total:
                    parts.append(f"--- {name} --- (omitted for space)")
                    continue
                used += len(b)
                parts.append(f"--- {name} ---\n{b}")
            user += (_REPO_DEV_PARENT_BLOCK.format(note=(f"; {base_note}" if base_note else ""))
                     + "\n\n".join(parts))
        if error:
            already = ", ".join(self.last_files) or "(none)"
            user += _REPO_DEV_REPAIR_BLOCK.format(already=already) + error[:4000]
        # A fresh implement (not a repair) on a real repo runs THREE explicit, separately-traced phases —
        # each its own focused tool-loop + emit so the context stays small and the trace reads cleanly
        # (Developer · stages → plan → implement):
        #   1. STAGES (mandatory, unless the operator declared `eval.stages` or protected the manifest):
        #      declare the ordered eval pipeline (prep → train → …) around the operator's protected
        #      `score` cmd — hardcoding this node's train params / adding a data_prep stage where useful.
        #      The Developer knows the repo; the planner (Genesis) may not.
        #   2. PLAN: decompose the code changes into ATOMIC steps (C4 — bounds a non-converging model).
        #   3. IMPLEMENT: write the code, one bounded session per plan step (each step its own trace block).
        # A REPAIR (error set) OR a bare / __new__-constructed dev (unit tests, no `_editables`) skips
        # straight to a single bounded session — repair is already narrow; the toy dev has no repo to stage.
        is_fresh_repo = error is None and getattr(self, "_editables", None)
        from looplab.agents.agent import CompositeTools
        from looplab.tools.env_inspect import EnvInspectTools
        try:
            operator_stages: list = []
            declared: list = []
            manifest_protected = False
            if is_fresh_repo:
                # Skip the STAGES phase when the OPERATOR already declared an `eval.stages` pipeline the
                # engine will actually USE: _resolve_stages takes a VALID operator list verbatim (a
                # Developer manifest would be IGNORED) but falls through to the Developer manifest on an
                # invalid one — so gate on the SAME shared validation, not truthiness. Protecting
                # `looplab_stages.json` is the operator knob that disables Developer pipelines entirely:
                # skip the phase (its manifest could never materialize) instead of burning a full LLM
                # loop whose output workspace-materialization silently drops.
                ev0 = self._cmd_context()[0]
                if ev0.get("stages"):
                    from looplab.runtime.command_eval import validate_stages
                    operator_stages = validate_stages(ev0["stages"])[0] or []
                manifest_protected = SurfacePolicy(
                    None, self._protected, self._prefixes, protected_exact=True,
                    check_escapes=False).check("looplab_stages.json") is not None
                if operator_stages:
                    declared = operator_stages
                elif not manifest_protected:
                    # STAGES is the Developer's own sub-phase (its own trace band, via the phase
                    # stamped on its generations).
                    with tracing.operation("stages"):
                        declared = self._declare_stages_phase(idea, write, system) or []
                # Tell the implement sessions what pipeline ACTUALLY exists. The old prompt asserted
                # "your STAGES phase already declared a train stage" unconditionally — after a failed/
                # empty stages phase the model then wrote a score-only entrypoint that scored a stale
                # checkpoint (or crashed on a missing one) instead of training.
                _chain = " → ".join(str(s.get("name")) for s in declared)
                if operator_stages:
                    stage_note = (f"\nPIPELINE for this node (OPERATOR-declared, runs verbatim): "
                                  f"{_chain}. Implement the code those stages run.")
                elif declared:
                    stage_note = (f"\nPIPELINE for this node (declared by your STAGES phase): {_chain} "
                                  "→ score (operator cmd). Implement the code those stages run; the "
                                  "eval entrypoint only SCORES the artifacts the earlier stages produce.")
                else:
                    stage_note = ("\nNO pipeline stages are declared for this node"
                                  + (" (the operator protected looplab_stages.json)"
                                     if manifest_protected else "")
                                  + ": the operator's cmd runs ALONE as a single command. The code it "
                                  "runs must do ALL the work itself when invoked — train a FRESH model, "
                                  "then score it and print the metric (never read a pre-existing "
                                  "checkpoint or a static results file).")
                user += stage_note
            messages = [{"role": "system", "content": system}, {"role": "user", "content": user}]
            # Compose the write/edit tools with read-only ENVIRONMENT INTROSPECTION (pkg_info / py_api /
            # read_installed / grep_installed) so the Developer grounds generated code in the ACTUAL
            # installed API/version instead of guessing (the precision='16-mixed'-on-Lightning-1.5 class).
            tools = CompositeTools([write, EnvInspectTools()] + self._scout_tools(write))
            if is_fresh_repo:
                # PLAN is the Developer's second sub-phase (its own trace band). IMPLEMENT runs under
                # the orchestrator's "implement" span (so its generations band there, and non-repo
                # developers keep that band unchanged).
                steps = []
                if getattr(self, "_plan_decompose", False):
                    with tracing.operation("plan"):
                        steps = self._propose_plan(system, idea, write)
                if len(steps) >= getattr(self, "_plan_min_steps", 2):
                    for i, step in enumerate(steps, 1):
                        self._run_step(idea, step, i, len(steps), write, system,
                                       stage_note=stage_note)  # a step error can't abort the plan
                else:
                    # single-session implement is TERMINAL (evaluation reads no brief) → consume the
                    # briefs + read-cache, but no wasted summary call (handoff=False).
                    run_phase(self.client, tools, messages, self._emit_spec(),
                              label="Developer·implement", handoff=False,
                              finalize=lambda a: (a or {}).get("summary", ""),
                              fallback=lambda m: "", **self._session_opts())
            else:
                # repair / toy single session — terminal, so no summary (and repair isn't in a scope
                # anyway when it runs inline during eval; the debug-operator repair gets an empty ledger).
                run_phase(self.client, tools, messages, self._emit_spec(),
                          label=("Developer·repair" if error else "Developer·implement"), handoff=False,
                          finalize=lambda a: (a or {}).get("summary", ""),
                          fallback=lambda m: "", **self._session_opts())
        except Exception as e:  # noqa: BLE001 - never crash the engine on a developer hiccup
            self.last_files = dict(write.files)
            self.last_deleted = list(write.deleted)
            return f"(developer error: {e})"
        self.last_files = dict(write.files)
        self.last_deleted = list(write.deleted)
        return ""

    def implement(self, idea: Idea) -> str:
        return self._run(idea)

    def implement_from(self, idea: Idea, parent) -> str:
        """Improve/refine: start from the PARENT node's solution and patch it (see _run(base=...)).
        Falls back to a from-scratch implement when the parent carries no files AND no deletions
        (e.g. seeded rows)."""
        files = dict(getattr(parent, "files", {}) or {})
        deleted = list(getattr(parent, "deleted", []) or [])
        if not files and not deleted:
            return self._run(idea)
        note = f"parent experiment #{getattr(parent, 'id', '?')}, metric={getattr(parent, 'metric', None)}"
        return self._run(idea, base=files, base_note=note, base_deleted=deleted)

    def repair(self, idea: Idea, code: str, error: str) -> str:
        return self._run(idea, error=error)

    def repair_from(self, idea: Idea, node, error: str) -> str:
        """Repair seeded from the FAILING NODE's OWN files (not the shared developer's `last_files`,
        which holds whatever node it built last — almost never this one). Falls back to the legacy
        last_files carry only when the node has no files (single-file / non-repo)."""
        files = dict(getattr(node, "files", {}) or {})
        deleted = list(getattr(node, "deleted", []) or [])
        if not files and not deleted:
            return self._run(idea, error=error)
        return self._run(idea, error=error, base=files, base_deleted=deleted)


class LLMOnboarder:
    """Phase 3 onboarder: the operator gives the framework's command; the Developer writes a
    metric `adapter` (read_metric(workdir)->float) that extracts the metric from whatever
    tracker/logs the run produced (TensorBoard / MLflow / metrics file / stdout). Returns a
    proposal that a human ratifies (then it's frozen + protected). Writing the adapter code
    is the Developer's job — onboarding reuses the same role, not a bespoke agent."""

    _SYS = ("You write a single Python module that reads the FINAL evaluation metric a "
            "training run produced. Output ONLY one ```python``` block defining "
            "`read_metric(workdir: str) -> float`.")

    def __init__(self, client, repo_path, goal, direction, command, timeout):
        self.client = client
        self.repo_path = repo_path
        self.goal = goal
        self.direction = direction
        self.command = command
        self.timeout = timeout

    def _context(self) -> tuple[str, str]:
        """Repo listing + the contents of a few small text files (the entrypoint, configs)
        so the Developer can see the actual metric shape it must read."""
        from pathlib import Path as _P
        root = _P(self.repo_path)
        _skip = {".git", "__pycache__", ".venv", "node_modules", ".mypy_cache", ".pytest_cache"}
        files = [p for p in root.rglob("*")
                 if p.is_file() and _skip.isdisjoint(p.parts)]
        listing = "\n".join(str(p.relative_to(root)) for p in files[:60])
        snippets, exts = [], (".py", ".json", ".yaml", ".yml", ".cfg", ".toml", ".txt")
        for p in files:
            if p.suffix in exts and p.stat().st_size < 4000:
                snippets.append(f"--- {p.relative_to(root)} ---\n"
                                + p.read_text(encoding="utf-8", errors="replace")[:2000])
            if len(snippets) >= 6:
                break
        return listing, "\n\n".join(snippets)

    def __call__(self) -> dict:
        from looplab.core.parse import extract_code
        cmd = " ".join(self.command) or "(the project's training command)"
        listing, snippets = self._context()
        user = (f"Repository files:\n{listing}\n\nKey file contents:\n{snippets}\n\n"
                f"The training command `{cmd}` runs in the work directory. Goal: {self.goal} "
                f"({self.direction}imize). Write `read_metric(workdir)` that, AFTER the run, "
                "returns the final metric by reading what the framework wrote FOR THIS RUN (match the "
                "metric key/format you see in the files above — e.g. a JSON like "
                '{"metric": <float>}). Read ONLY the CURRENT run\'s freshly-written output; NEVER read a '
                "pre-existing/committed results file or a prior run's checkpoint (e.g. results_last.csv is "
                "a PRIOR run's output, not this run's score). Prefer stdlib; if you use an optional tracker lib "
                "(tensorboard/mlflow), import it INSIDE a try/except and fall back. Return a "
                "float; on any problem return a clearly-bad value so the run is not rewarded.")
        try:
            code = extract_code(self.client.complete_text(
                [{"role": "system", "content": self._SYS}, {"role": "user", "content": user}]))
        except Exception as e:  # noqa: BLE001 — propose a stub; human will reject/fix
            code = f"def read_metric(workdir):\n    raise RuntimeError({str(e)!r})\n"
        return {
            "eval_spec": {"command": list(self.command),
                          "metric": {"kind": "adapter", "path": "LOOPLAB_adapter.py"},
                          "params_style": "none", "timeout": self.timeout},
            "adapter_files": {"LOOPLAB_adapter.py": code},
            "goal": self.goal,
        }
